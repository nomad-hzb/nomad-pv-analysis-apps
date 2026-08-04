import io

import data_manager
import openpyxl
import pytest
from data_manager import (
    LABOR_ROLES,
    BatchTotal,
    CostReference,
    LCCDataManager,
    MaterialCostRow,
    MaterialRow,
    ProcessCostRow,
    ProcessStepRow,
    compute_batch_totals,
    compute_labor_cost,
    compute_material_cost_rows,
    compute_process_cost_rows,
    discover_entry_types,
    extract_material_rows,
    extract_process_step_rows,
    fetch_process_entries,
    fetch_samples_per_batch,
    flatten_entry,
    load_default_cost_reference,
    parse_cost_reference_workbook,
)
from excel_export import (
    LABOR_HEADERS,
    MATERIALS_HEADERS,
    PROCESSES_HEADERS,
    SUMMARY_HEADERS,
    build_cost_reference_template,
    build_cost_reference_template_from_data,
    build_workbook,
)

# ---------------------------------------------------------------------------
# flatten_entry
# ---------------------------------------------------------------------------


def test_flatten_entry_nested_dict():
    flat = flatten_entry({"annealing": {"temperature": 120, "time": 30}})
    assert flat == {"annealing.temperature": 120, "annealing.time": 30}


def test_flatten_entry_single_item_list_unwrapped():
    flat = flatten_entry({"layer": [{"layer_material_name": "PbI2"}]})
    assert flat == {"layer.layer_material_name": "PbI2"}


def test_flatten_entry_multi_item_list_indexed():
    flat = flatten_entry({"recipe_steps": [{"time": 30}, {"time": 45}]})
    assert flat == {"recipe_steps.1.time": 30, "recipe_steps.2.time": 45}


def test_flatten_entry_empty_list_dropped():
    flat = flatten_entry({"layer": []})
    assert flat == {}


# ---------------------------------------------------------------------------
# extract_process_step_rows
# ---------------------------------------------------------------------------


def test_extract_process_step_rows_single_step_with_top_level_time():
    process_data = {
        "name": "ALD process",
        "positon_in_experimental_plan": 2,
        "location": "HyALD",
        "time": 1800,
        "rate": 0.1,
    }
    rows = extract_process_step_rows(process_data, "HySprint_ALD", "Batch1", 4, ["S1", "S2"])
    assert len(rows) == 1
    row = rows[0]
    assert row.step_index == 0
    assert row.duration_value == 1800
    assert row.rate_value == 0.1
    assert row.num_samples_covered == 4
    assert row.position_in_plan == 2
    assert row.location == "HyALD"


def test_extract_process_step_rows_expands_recipe_steps():
    process_data = {
        "positon_in_experimental_plan": 1,
        "recipe_steps": [
            {"time": 30, "speed": 1000},
            {"time": 45, "speed": 2000},
        ],
    }
    rows = extract_process_step_rows(process_data, "HySprint_SpinCoating", "Batch1", 1, ["S1"])
    assert len(rows) == 2
    assert all(row.step_label == "recipe_steps" for row in rows)
    assert {row.duration_value for row in rows} == {30, 45}
    assert {row.rate_value for row in rows} == {1000, 2000}


def test_extract_process_step_rows_skips_non_step_lists():
    process_data = {
        "positon_in_experimental_plan": 1,
        "solution": [
            {
                "solution_details": {
                    "solute": [
                        {"chemical_2": {"name": "PbI2"}, "concentration_mol": 1.4},
                        {"chemical_2": {"name": "MAI"}, "concentration_mol": 1.4},
                    ]
                }
            }
        ],
    }
    rows = extract_process_step_rows(process_data, "HySprint_SpinCoating", "Batch1", 1, ["S1"])
    assert all(not row.step_label.startswith("solution") for row in rows)


def test_extract_process_step_rows_missing_location_defaults_empty():
    process_data = {"positon_in_experimental_plan": 1, "method": "Cleaning"}
    rows = extract_process_step_rows(process_data, "HySprint_Cleaning", "Batch1", 1, ["S1"])
    assert rows[0].location == ""


# ---------------------------------------------------------------------------
# extract_material_rows
# ---------------------------------------------------------------------------


def test_extract_material_rows_solution_solute_and_solvent():
    process_data = {
        "solution": [
            {
                "solution_details": {
                    "solute": [{"chemical_2": {"name": "PbI2"}, "concentration_mol": 1.4}],
                    "solvent": [{"chemical_2": {"name": "DMF"}, "chemical_volume": 500}],
                }
            }
        ]
    }
    rows = extract_material_rows(process_data, "HySprint_SpinCoating", "Batch1", 2, ["S1", "S2"])
    names_roles = {(row.material_name, row.role) for row in rows}
    assert names_roles == {("PbI2", "solute"), ("DMF", "solvent")}

    pbi2 = next(row for row in rows if row.material_name == "PbI2")
    assert pbi2.molar_mass_g_per_mol == pytest.approx(461.01)
    assert pbi2.price_per_gram_est is not None
    assert pbi2.cas_number == "10101-63-0"
    assert pbi2.num_samples_covered == 2


def test_extract_material_rows_unknown_chemical_has_no_reference():
    process_data = {"layer": [{"layer_material_name": "SomeNovelCompound"}]}
    rows = extract_material_rows(process_data, "HySprint_Evaporation", "Batch1", 1, ["S1"])
    assert len(rows) == 1
    assert rows[0].molar_mass_g_per_mol is None
    assert rows[0].price_per_gram_est is None
    assert rows[0].cas_number is None


def test_extract_material_rows_evaporant():
    process_data = {"organic_evaporation": [{"chemical_2": {"name": "C60"}, "thickness": 20}]}
    rows = extract_material_rows(process_data, "HySprint_Evaporation", "Batch1", 1, ["S1"])
    assert len(rows) == 1
    assert rows[0].role == "evaporant"
    assert rows[0].quantity_value == 20


def test_extract_material_rows_prefers_real_nomad_cas_over_static_table():
    process_data = {
        "solution": [
            {
                "solution_details": {
                    "solute": [
                        {
                            "chemical_2": {"name": "PbI2", "cas_number": "REAL-CAS-FROM-NOMAD"},
                            "concentration_mol": 1.0,
                        }
                    ]
                }
            }
        ]
    }
    rows = extract_material_rows(process_data, "HySprint_SpinCoating", "Batch1", 1, ["S1"])
    assert rows[0].cas_number == "REAL-CAS-FROM-NOMAD"


# ---------------------------------------------------------------------------
# NOMAD fetch functions (mocked)
# ---------------------------------------------------------------------------


def test_fetch_samples_per_batch(mocker):
    mock_get_ids = mocker.patch("hysprint_utils.api_calls.get_ids_in_batch")
    mock_get_ids.side_effect = lambda url, token, batch_ids: [
        f"{batch_ids[0]}_S1",
        f"{batch_ids[0]}_S2",
    ]

    result = fetch_samples_per_batch("http://x", "tok", ["Batch1", "Batch2"])

    assert result == {
        "Batch1": ["Batch1_S1", "Batch1_S2"],
        "Batch2": ["Batch2_S1", "Batch2_S2"],
    }
    assert mock_get_ids.call_count == 2


def test_fetch_samples_per_batch_skips_batch_that_raises_assertion_error(mocker):
    mock_get_ids = mocker.patch("hysprint_utils.api_calls.get_ids_in_batch")

    def side_effect(url, token, batch_ids):
        if batch_ids == ["BadBatch"]:
            raise AssertionError
        return [f"{batch_ids[0]}_S1"]

    mock_get_ids.side_effect = side_effect

    result = fetch_samples_per_batch("http://x", "tok", ["GoodBatch", "BadBatch"])

    assert result == {"GoodBatch": ["GoodBatch_S1"]}


def test_fetch_samples_per_batch_skips_batch_that_raises_key_error(mocker):
    """Confirmed against a full scan of every real batch on a live instance:
    malformed/legacy batch data can make get_ids_in_batch raise KeyError
    (an "entities" item missing "lab_id"), not just AssertionError."""
    mock_get_ids = mocker.patch("hysprint_utils.api_calls.get_ids_in_batch")

    def side_effect(url, token, batch_ids):
        if batch_ids == ["MalformedBatch"]:
            raise KeyError("lab_id")
        return [f"{batch_ids[0]}_S1"]

    mock_get_ids.side_effect = side_effect

    result = fetch_samples_per_batch("http://x", "tok", ["GoodBatch", "MalformedBatch"])

    assert result == {"GoodBatch": ["GoodBatch_S1"]}


def test_fetch_process_entries_keeps_full_samples_list_and_filters_unpositioned(mocker):
    entry_id_response = mocker.Mock()
    entry_id_response.json.return_value = {"data": [{"entry_id": "e1"}]}
    entry_id_response.raise_for_status.return_value = None

    process_response = mocker.Mock()
    process_response.json.return_value = {
        "data": [
            {
                "archive": {
                    "data": {
                        "positon_in_experimental_plan": 2,
                        "samples": [{"lab_id": "S1"}, {"lab_id": "S2"}, {"lab_id": "S3"}],
                    },
                    "metadata": {"entry_type": "HySprint_Evaporation"},
                }
            },
            {
                "archive": {
                    "data": {"positon_in_experimental_plan": 1, "samples": [{"lab_id": "S1"}]},
                    "metadata": {"entry_type": "HySprint_SpinCoating"},
                }
            },
            {"archive": {"data": {"samples": [{"lab_id": "S1"}]}, "metadata": {}}},
        ]
    }
    process_response.raise_for_status.return_value = None

    mock_post = mocker.patch("requests.post", side_effect=[entry_id_response, process_response])

    entries = fetch_process_entries("http://x", "tok", ["S1", "S2", "S3"])

    assert mock_post.call_count == 2
    assert [entry_type for _data, entry_type in entries] == [
        "HySprint_SpinCoating",
        "HySprint_Evaporation",
    ]
    evaporation_data = entries[1][0]
    assert len(evaporation_data["samples"]) == 3


def test_discover_entry_types_excludes_structural_types(mocker):
    entry_id_response = mocker.Mock()
    entry_id_response.json.return_value = {"data": [{"entry_id": "e1"}]}
    entry_id_response.raise_for_status.return_value = None

    aggregation_response = mocker.Mock()
    aggregation_response.json.return_value = {
        "aggregations": {
            "entry_type_agg": {
                "terms": {
                    "data": [
                        {"value": "HySprint_JVmeasurement"},
                        {"value": "HySprint_Evaporation"},
                        {"value": "HySprint_Batch"},
                        {"value": "HySprint_Sample"},
                    ]
                }
            }
        }
    }
    aggregation_response.raise_for_status.return_value = None

    mocker.patch("requests.post", side_effect=[entry_id_response, aggregation_response])

    result = discover_entry_types("http://x", "tok", ["S1"])

    assert result == ["HySprint_Evaporation", "HySprint_JVmeasurement"]


def test_discover_entry_types_no_entries_returns_empty(mocker):
    entry_id_response = mocker.Mock()
    entry_id_response.json.return_value = {"data": []}
    entry_id_response.raise_for_status.return_value = None
    mocker.patch("requests.post", return_value=entry_id_response)

    assert discover_entry_types("http://x", "tok", ["S1"]) == []


# ---------------------------------------------------------------------------
# LCCDataManager.load_batches (integration over the extraction pipeline)
# ---------------------------------------------------------------------------


def test_load_batches_populates_rows_and_batch_sample_counts(mocker):
    mocker.patch.object(
        data_manager,
        "fetch_samples_per_batch",
        return_value={"Batch1": ["S1", "S2"]},
    )
    mocker.patch.object(
        data_manager,
        "fetch_process_entries",
        return_value=[
            (
                {
                    "positon_in_experimental_plan": 1,
                    "location": "HyVapBox",
                    "time": 60,
                    "samples": [{"lab_id": "S1"}, {"lab_id": "S2"}],
                    "layer": [{"layer_material_name": "PbI2"}],
                },
                "HySprint_Evaporation",
            )
        ],
    )

    manager = LCCDataManager()
    manager.load_batches("http://x", "tok", ["Batch1"])

    assert manager.batch_sample_counts == {"Batch1": 2}
    assert manager.all_sample_ids == ["S1", "S2"]
    assert len(manager.process_rows) == 1
    assert manager.process_rows[0].num_samples_covered == 2
    assert manager.process_rows[0].location == "HyVapBox"
    assert len(manager.material_rows) == 1
    assert manager.material_rows[0].material_name == "PbI2"
    assert manager.has_data is True


# ---------------------------------------------------------------------------
# Cost computation - process rows
# ---------------------------------------------------------------------------


def test_compute_process_cost_rows_aggregates_by_type_and_separates_costs():
    process_rows = [
        ProcessStepRow(
            batch_id="Batch1",
            process_type="HySprint_SpinCoating",
            location="HySpinBox",
            step_index=0,
            num_samples_covered=3,
        ),
        ProcessStepRow(
            batch_id="Batch1",
            process_type="HySprint_SpinCoating",
            location="HySpinBox",
            step_index=1,
            num_samples_covered=3,
        ),
    ]
    reference = CostReference(
        process_costs={"HySprint_SpinCoating": {"cost_est": 10.0, "verified": True, "notes": ""}},
        location_costs={"HySpinBox": {"cost_est": 25.0, "verified": True, "notes": ""}},
    )

    rows = compute_process_cost_rows(process_rows, {"Batch1": 3}, reference)

    assert len(rows) == 1
    row = rows[0]
    assert row.step_count == 2
    assert row.process_cost == 10.0
    assert row.equipment_cost == 25.0
    assert row.total_cost == 35.0  # separate figures, summed only for total
    assert row.num_samples == 3


def test_compute_process_cost_rows_no_reference_leaves_costs_none():
    process_rows = [ProcessStepRow(batch_id="Batch1", process_type="X", step_index=0)]
    rows = compute_process_cost_rows(process_rows, {"Batch1": 1}, None)
    assert rows[0].process_cost is None
    assert rows[0].equipment_cost is None
    assert rows[0].total_cost is None


def test_compute_process_cost_rows_sums_multiple_locations():
    process_rows = [
        ProcessStepRow(batch_id="Batch1", process_type="X", location="BoxA", step_index=0),
        ProcessStepRow(batch_id="Batch1", process_type="X", location="BoxB", step_index=1),
    ]
    reference = CostReference(
        location_costs={
            "BoxA": {"cost_est": 5.0, "verified": True, "notes": ""},
            "BoxB": {"cost_est": 7.0, "verified": False, "notes": ""},
        }
    )
    rows = compute_process_cost_rows(process_rows, {"Batch1": 1}, reference)
    assert rows[0].equipment_cost == 12.0
    assert rows[0].equipment_cost_verified is False  # not all matched locations verified


# ---------------------------------------------------------------------------
# Cost computation - material rows
# ---------------------------------------------------------------------------


def test_compute_material_cost_rows_uses_average_quantity_times_usage_count():
    material_rows = [
        MaterialRow(batch_id="Batch1", process_type="X", material_name="PbI2", role="solute"),
        MaterialRow(batch_id="Batch1", process_type="Y", material_name="PbI2", role="solute"),
    ]
    reference = CostReference(
        material_prices={
            "PbI2": {
                "cas_number": "10101-63-0",
                "price_per_gram_est": 3.0,
                "average_quantity_grams": 0.5,
                "verified": True,
                "notes": "",
            }
        }
    )

    rows = compute_material_cost_rows(material_rows, reference)

    assert len(rows) == 1
    row = rows[0]
    assert row.usage_count == 2
    assert row.quantity_grams == pytest.approx(1.0)  # 0.5 x 2 uses
    assert row.quantity_source == "average (reference)"
    assert row.total_cost == pytest.approx(3.0)


def test_compute_material_cost_rows_unknown_quantity_when_no_average_in_reference():
    material_rows = [
        MaterialRow(batch_id="Batch1", process_type="X", material_name="PbI2", role="solute")
    ]
    rows = compute_material_cost_rows(material_rows, None)
    assert rows[0].quantity_grams is None
    assert rows[0].quantity_source == "unknown - not in cost reference yet"
    assert rows[0].total_cost is None


# ---------------------------------------------------------------------------
# Cost computation - labor and batch totals
# ---------------------------------------------------------------------------


def test_compute_labor_cost_multiplies_hours_by_rate():
    reference = CostReference(labor_rates={"Postdoc": {"hourly_rate_est": 45.0, "verified": True}})
    cost, verified = compute_labor_cost("Postdoc", 2.0, reference)
    assert cost == 90.0
    assert verified is True


def test_compute_labor_cost_none_when_no_hours():
    reference = CostReference(labor_rates={"Postdoc": {"hourly_rate_est": 45.0, "verified": True}})
    cost, verified = compute_labor_cost("Postdoc", 0.0, reference)
    assert cost is None
    assert verified is False


def test_compute_labor_cost_none_when_role_not_in_reference():
    cost, verified = compute_labor_cost("Postdoc", 2.0, None)
    assert cost is None
    assert verified is False


def test_compute_batch_totals_sums_and_counts_unverified():
    process_rows = [
        ProcessCostRow(
            batch_id="Batch1",
            process_type="X",
            step_count=1,
            locations=[],
            process_cost=10.0,
            process_cost_verified=True,
            equipment_cost=None,
            equipment_cost_verified=False,
            num_samples=2,
        )
    ]
    material_rows = [
        MaterialCostRow(
            batch_id="Batch1",
            material_name="PbI2",
            roles=["solute"],
            usage_count=1,
            cas_number=None,
            quantity_grams=1.0,
            quantity_source="average (reference)",
            price_per_gram=3.0,
            total_cost=3.0,
            verified=False,
            notes="",
        )
    ]
    labor_costs = {"Batch1": (90.0, True)}

    totals = compute_batch_totals(process_rows, material_rows, {"Batch1": 2}, labor_costs)

    assert len(totals) == 1
    total = totals[0]
    assert total.process_total == 10.0
    assert total.material_total == 3.0
    assert total.labor_total == 90.0
    assert total.grand_total == pytest.approx(103.0)
    assert total.per_sample == pytest.approx(51.5)
    assert total.unverified_count == 1  # only the unverified material row


def test_batch_total_per_sample_none_when_zero_samples():
    total = BatchTotal(
        batch_id="B",
        num_samples=0,
        material_total=0.0,
        process_total=0.0,
        equipment_total=0.0,
        labor_total=0.0,
        unverified_count=0,
    )
    assert total.per_sample is None


# ---------------------------------------------------------------------------
# Cost reference parsing
# ---------------------------------------------------------------------------


def test_parse_cost_reference_workbook_reads_all_sheets():
    workbook = build_cost_reference_template()
    materials_ws = workbook["Materials"]
    pbi2_row = next(
        r
        for r in range(2, materials_ws.max_row + 1)
        if materials_ws.cell(row=r, column=1).value == "PbI2"
    )
    materials_ws.cell(row=pbi2_row, column=3, value=500.0)  # Price
    materials_ws.cell(row=pbi2_row, column=4, value=250.0)  # Grams_on_Bottle
    materials_ws.cell(row=pbi2_row, column=6, value=0.5)  # Average_Quantity_Grams
    materials_ws.cell(row=pbi2_row, column=7, value=True)  # Verified

    processes_ws = workbook["Processes"]
    processes_ws.cell(row=2, column=1, value="HySprint_SpinCoating")
    processes_ws.cell(row=2, column=2, value=10.0)
    processes_ws.cell(row=2, column=3, value=True)

    labor_ws = workbook["Labor"]
    labor_ws.cell(row=2, column=2, value=45.0)  # PhD Researcher rate
    labor_ws.cell(row=2, column=3, value=True)

    capital_ws = workbook["Capital_Overhead_Disposal"]
    capital_ws.cell(row=2, column=5, value=15.0)  # Cost for "Cleanroom / lab rent"
    capital_ws.cell(row=2, column=6, value=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    reference = parse_cost_reference_workbook(buffer.getvalue())

    assert reference.material_prices["PbI2"]["price_per_gram_est"] == pytest.approx(2.0)
    assert reference.material_prices["PbI2"]["average_quantity_grams"] == pytest.approx(0.5)
    assert reference.material_prices["PbI2"]["verified"] is True

    assert reference.process_costs["HySprint_SpinCoating"]["cost_est"] == 10.0
    assert reference.process_costs["HySprint_SpinCoating"]["verified"] is True

    assert reference.labor_rates["PhD Researcher"]["hourly_rate_est"] == 45.0

    assert reference.overhead_costs["Cleanroom / lab rent"]["cost_est"] == 15.0


def test_parse_cost_reference_workbook_populates_location_costs():
    workbook = build_cost_reference_template()
    capital_ws = workbook["Capital_Overhead_Disposal"]
    # First KNOWN_LOCATIONS row after the rent row.
    capital_ws.cell(row=3, column=5, value=25.0)
    capital_ws.cell(row=3, column=6, value=True)
    location = capital_ws.cell(row=3, column=2).value

    buffer = io.BytesIO()
    workbook.save(buffer)
    reference = parse_cost_reference_workbook(buffer.getvalue())

    assert reference.location_costs[location]["cost_est"] == 25.0


# ---------------------------------------------------------------------------
# Automatic single-source-of-truth cost reference file
# ---------------------------------------------------------------------------


def test_load_default_cost_reference_missing_file_returns_none(tmp_path):
    missing_path = tmp_path / "does_not_exist.xlsx"
    assert load_default_cost_reference(path=missing_path) is None


def test_load_default_cost_reference_reads_existing_file(tmp_path):
    workbook = build_cost_reference_template()
    file_path = tmp_path / "cost_reference.xlsx"
    workbook.save(file_path)

    reference = load_default_cost_reference(path=file_path)

    assert reference is not None
    assert "PbI2" in reference.material_prices
    assert reference.material_prices["PbI2"]["price_per_gram_est"] == pytest.approx(3.0)


def test_load_default_cost_reference_corrupt_file_returns_none(tmp_path):
    file_path = tmp_path / "cost_reference.xlsx"
    file_path.write_bytes(b"not a real xlsx file")

    assert load_default_cost_reference(path=file_path) is None


# ---------------------------------------------------------------------------
# excel_export.build_workbook (per-batch report, literal values)
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_cost_rows():
    process_rows = [
        ProcessCostRow(
            batch_id="Batch1",
            process_type="HySprint_SpinCoating",
            step_count=2,
            locations=["HySpinBox"],
            process_cost=10.0,
            process_cost_verified=True,
            equipment_cost=25.0,
            equipment_cost_verified=True,
            num_samples=3,
        )
    ]
    material_rows = [
        MaterialCostRow(
            batch_id="Batch1",
            material_name="PbI2",
            roles=["solute"],
            usage_count=1,
            cas_number="10101-63-0",
            quantity_grams=0.5,
            quantity_source="average (reference)",
            price_per_gram=3.0,
            total_cost=1.5,
            verified=True,
            notes="",
        )
    ]
    labor_selections = {"Batch1": ("Postdoc", 2.0)}
    cost_reference = CostReference(
        labor_rates={"Postdoc": {"hourly_rate_est": 45.0, "verified": True}}
    )
    labor_costs = {
        batch_id: compute_labor_cost(role, hours, cost_reference)
        for batch_id, (role, hours) in labor_selections.items()
    }
    batch_totals = compute_batch_totals(process_rows, material_rows, {"Batch1": 3}, labor_costs)
    return process_rows, material_rows, batch_totals, labor_selections, cost_reference


def test_build_workbook_sheet_names(sample_cost_rows):
    workbook = build_workbook(*sample_cost_rows)
    assert workbook.sheetnames == ["Guide", "Processes", "Materials", "Labor", "Summary"]


def test_build_workbook_headers_match(sample_cost_rows):
    workbook = build_workbook(*sample_cost_rows)
    assert [c.value for c in workbook["Processes"][1]] == PROCESSES_HEADERS
    assert [c.value for c in workbook["Materials"][1]] == MATERIALS_HEADERS
    assert [c.value for c in workbook["Labor"][1]] == LABOR_HEADERS
    assert [c.value for c in workbook["Summary"][1]] == SUMMARY_HEADERS


def test_build_workbook_processes_row_has_separate_process_and_equipment_cost(sample_cost_rows):
    workbook = build_workbook(*sample_cost_rows)
    ws = workbook["Processes"]
    process_cost_col = PROCESSES_HEADERS.index("Process_Cost") + 1
    equipment_cost_col = PROCESSES_HEADERS.index("Equipment_Cost") + 1
    total_cost_col = PROCESSES_HEADERS.index("Total_Cost") + 1
    assert ws.cell(row=2, column=process_cost_col).value == 10.0
    assert ws.cell(row=2, column=equipment_cost_col).value == 25.0
    assert ws.cell(row=2, column=total_cost_col).value == 35.0


def test_build_workbook_cells_are_literal_values_not_formulas(sample_cost_rows):
    """The per-batch report is a read-only computed snapshot - no formulas."""
    workbook = build_workbook(*sample_cost_rows)
    for sheet_name in ("Processes", "Materials", "Labor", "Summary"):
        ws = workbook[sheet_name]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("="), f"{sheet_name}!{cell.coordinate}"


def test_build_workbook_summary_includes_total_row(sample_cost_rows):
    workbook = build_workbook(*sample_cost_rows)
    ws = workbook["Summary"]
    batch_id_col = SUMMARY_HEADERS.index("Batch_ID") + 1
    last_row_value = ws.cell(row=ws.max_row, column=batch_id_col).value
    assert last_row_value == "TOTAL (all selected)"


def test_build_workbook_round_trips_through_bytes(sample_cost_rows):
    workbook = build_workbook(*sample_cost_rows)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    reloaded = openpyxl.load_workbook(buffer)
    assert reloaded.sheetnames == workbook.sheetnames
    assert reloaded["Summary"].cell(row=2, column=1).value == "Batch1"


# ---------------------------------------------------------------------------
# Cost reference template builders (admin single source of truth)
# ---------------------------------------------------------------------------


def test_build_cost_reference_template_sheet_names():
    workbook = build_cost_reference_template()
    assert workbook.sheetnames == [
        "Guide",
        "Materials",
        "Processes",
        "Labor",
        "Capital_Overhead_Disposal",
    ]


def test_build_cost_reference_template_materials_seeded_with_cas_and_price():
    workbook = build_cost_reference_template()
    ws = workbook["Materials"]
    pbi2_row = next(r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == "PbI2")
    assert ws.cell(row=pbi2_row, column=2).value == "10101-63-0"  # CAS_Number
    assert ws.cell(row=pbi2_row, column=3).value == pytest.approx(3.0)  # Price
    assert ws.cell(row=pbi2_row, column=4).value == 1  # Grams_on_Bottle
    assert ws.cell(row=pbi2_row, column=6).value is None  # Average_Quantity_Grams never guessed


def test_build_cost_reference_template_labor_has_four_fixed_roles():
    workbook = build_cost_reference_template()
    ws = workbook["Labor"]
    roles = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert roles == LABOR_ROLES


def test_build_cost_reference_template_capital_seeded_with_all_known_locations():
    workbook = build_cost_reference_template()
    ws = workbook["Capital_Overhead_Disposal"]
    items = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "Cleanroom / lab rent" in items
    assert "Equipment depreciation - HyVapBox" in items
    assert "Equipment depreciation - HyWeighBox" in items


def test_build_cost_reference_template_from_data_uses_real_names():
    workbook = build_cost_reference_template_from_data(
        process_types=["HySprint_Evaporation"],
        material_names_with_cas={"PbI2": None},
        batch_count=1,
    )

    materials_ws = workbook["Materials"]
    names = {materials_ws.cell(row=r, column=1).value for r in range(2, materials_ws.max_row + 1)}
    assert "PbI2" in names  # real, from the scanned batch
    assert "MAI" in names  # static-only fallback, offered anyway

    processes_ws = workbook["Processes"]
    process_types = {
        processes_ws.cell(row=r, column=1).value for r in range(2, processes_ws.max_row + 1)
    }
    assert "HySprint_Evaporation" in process_types

    capital_ws = workbook["Capital_Overhead_Disposal"]
    items = {capital_ws.cell(row=r, column=1).value for r in range(2, capital_ws.max_row + 1)}
    assert "Equipment depreciation - HyVapBox" in items  # from KNOWN_LOCATIONS, not batch data


def test_build_cost_reference_template_from_data_includes_extra_schema_types():
    workbook = build_cost_reference_template_from_data(
        process_types=["HySprint_Evaporation"],
        material_names_with_cas={},
        batch_count=1,
        extra_schema_types=["HySprint_JVmeasurement"],
    )
    processes_ws = workbook["Processes"]
    process_types = {
        processes_ws.cell(row=r, column=1).value for r in range(2, processes_ws.max_row + 1)
    }
    assert "HySprint_JVmeasurement" in process_types


def test_reference_template_round_trips_through_parser():
    workbook = build_cost_reference_template_from_data(
        process_types=["HySprint_Evaporation"],
        material_names_with_cas={"PbI2": "10101-63-0"},
        batch_count=1,
    )
    buffer = io.BytesIO()
    workbook.save(buffer)

    reference = parse_cost_reference_workbook(buffer.getvalue())

    assert "PbI2" in reference.material_prices
    assert reference.material_prices["PbI2"]["cas_number"] == "10101-63-0"
    assert set(reference.labor_rates) == set(LABOR_ROLES)
    assert "Cleanroom / lab rent" in reference.overhead_costs

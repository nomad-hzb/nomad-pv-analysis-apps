"""
excel_export.py
----------------
Builds the openpyxl Life Cycle Costing workbook from an LCCDataManager's
extracted rows. No widget imports - this module only produces an
openpyxl.Workbook; gui_components.py is responsible for turning that into
a browser download.
"""

from __future__ import annotations

import logging

from data_manager import (
    CHEMICAL_REFERENCE,
    CHEMICAL_REFERENCE_NOTE,
    KNOWN_LOCATIONS,
    LABOR_ROLES,
    CostReference,
    LCCDataManager,
)
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

MAX_DATA_ROW = 2000

HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
VERIFIED_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

COST_CATEGORY_OPTIONS = ["Capital", "Maintenance", "Overhead", "End-of-Life"]
ALLOCATION_METHOD_OPTIONS = ["per batch", "per sample", "per hour"]


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _col(headers: list[str], name: str) -> int:
    """1-based column index of `name` in `headers` - looked up by name so a
    sheet can gain/reorder columns without every formula needing a manual
    letter recount (a real source of bugs once a sheet has 15+ columns)."""
    return headers.index(name) + 1


def _letter(headers: list[str], name: str) -> str:
    return get_column_letter(_col(headers, name))


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _add_bool_dropdown(ws: Worksheet, col_letter: str, last_row: int = MAX_DATA_ROW) -> None:
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv.error = "Please select TRUE or FALSE"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def _add_list_dropdown(
    ws: Worksheet, col_letter: str, options: list[str], last_row: int = MAX_DATA_ROW
) -> None:
    options_str = ",".join(options)
    dv = DataValidation(type="list", formula1=f'"{options_str}"', allow_blank=True)
    dv.error = "Please select from the dropdown list"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def _highlight_verified_column(
    ws: Worksheet, verified_col_letter: str, last_row: int = MAX_DATA_ROW
) -> None:
    """Red for FALSE (unverified - the app's live "show me unverified cost
    values" requirement), green for TRUE (confirmed)."""
    cell_range = f"{verified_col_letter}2:{verified_col_letter}{last_row}"
    ws.conditional_formatting.add(
        cell_range, CellIsRule(operator="equal", formula=["FALSE"], fill=WARNING_FILL)
    )
    ws.conditional_formatting.add(
        cell_range, CellIsRule(operator="equal", formula=["TRUE"], fill=VERIFIED_FILL)
    )


def _autosize_columns(ws: Worksheet, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        letter = get_column_letter(col_index)
        ws.column_dimensions[letter].width = max(14, len(header) + 2)


# ---------------------------------------------------------------------------
# Guide sheet
# ---------------------------------------------------------------------------

_GUIDE_TEXT = [
    ("Life Cycle Costing (LCC) Calculator - Guide", True),
    ("", False),
    (
        "This workbook estimates the cost of each selected batch/experiment, "
        "broken down by the standard LCC cost categories:",
        False,
    ),
    ("  - Operation: materials/consumables (Materials sheet)", False),
    (
        "  - Personnel: labor time by role - PhD Researcher/Postdoc/Engineer/HiWi (Labor sheet)",
        False,
    ),
    (
        "  - Capital / Maintenance / Overhead / End-of-Life: equipment, rent, "
        "disposal (Capital_Overhead_Disposal sheet)",
        False,
    ),
    (
        "Process duration/rate/location are tracked on the Processes sheet "
        "for reference, but processes are not costed separately - their cost "
        "is already captured through the materials consumed and the labor "
        "time spent running them.",
        False,
    ),
    ("", False),
    ("Cost_Low / Cost_Est / Cost_High columns:", True),
    (
        "Fill in Cost_Est at minimum. Cost_Low/Cost_High are optional - if left "
        "blank, Cost_Est is used as both bounds automatically (no range).",
        False,
    ),
    ("", False),
    ("Verified column:", True),
    (
        "TRUE/FALSE dropdown. Set to TRUE once you've confirmed a cost figure "
        "against a real quote/invoice/rate - highlighted green. Unverified "
        "(FALSE) rows are highlighted red, and any batch with unverified "
        "line items is flagged on the Summary sheet.",
        False,
    ),
    ("", False),
    ("Labor sheet:", True),
    (
        "NOMAD does not record who ran a process or what role they hold, so "
        "labor cost cannot be extracted automatically. Instead, enter the "
        "estimated Hours worked per role (PhD Researcher/Postdoc/Engineer/"
        "HiWi) for this batch - the hourly rate is carried forward from the "
        "shared cost reference file's Labor sheet.",
        False,
    ),
    ("", False),
    ("Capital_Overhead_Disposal sheet - Location:", True),
    (
        "Equipment depreciation lines are tied to the physical box/tool "
        "(e.g. HySpinBox, HyVapBox). Note: raw NOMAD location text is "
        "inconsistent (typos/variants of the same box), so a depreciation "
        "cost only carries forward automatically when the batch's recorded "
        "location text exactly matches an Item name in the cost reference "
        "file - reconcile mismatches manually until location data is "
        "standardized at the source.",
        False,
    ),
    ("", False),
    ("Per-sample costs:", True),
    (
        "Num_Samples_Covered is how many samples a single material instance "
        "actually applies to (e.g. one evaporation run covering 16 samples "
        "costs the same per batch, and 1/16th per sample). The Summary "
        "sheet's per-sample figure is a batch average "
        "(Grand_Total / Num_Samples) - exact when all samples in a batch "
        "share identical processing, an approximation otherwise.",
        False,
    ),
    ("", False),
    ("Materials sheet - CAS_Number, Quantity_Grams and Price_per_Gram_Est:", True),
    (
        "CAS_Number is read directly from NOMAD when a chemist has actually "
        "populated it there (rare in practice - confirmed most chemical "
        "references in this system don't have it filled in), otherwise from "
        "the static reference table for well-known chemicals, otherwise "
        "blank. Quantity_Grams is intentionally left blank for solution-based "
        "materials (solutes/solvents) - NOMAD records concentration/volume, "
        "not a reliable gram amount, so auto-computing it risked being "
        "silently wrong. Price_per_Gram_Est is pre-filled for common "
        "perovskite-fab chemicals from a rough static reference table "
        "(general market knowledge, not a live supplier lookup) - always "
        "marked Verified=FALSE; please confirm.",
        False,
    ),
    ("", False),
    ("Where these costs come from:", True),
    (
        "Cost/Price/Rate and Verified values already filled in on this "
        "export were carried forward automatically from the shared admin-"
        "maintained cost reference file (cost_reference.xlsx, next to the "
        "app) - matched by Material_Name, Role (Labor), or Item (Capital_"
        "Overhead_Disposal). Only genuinely new line items start blank. To "
        "correct a value for everyone going forward, edit cost_reference.xlsx "
        "itself rather than this exported copy - an admin maintains that one "
        "file as the single source of truth.",
        False,
    ),
]


def _build_guide_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Guide"
    row = 1
    for text, bold in _GUIDE_TEXT:
        cell = ws.cell(row=row, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13 if row == 1 else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    ws.column_dimensions["A"].width = 110


# ---------------------------------------------------------------------------
# Processes sheet - informational only (duration/rate/location), no cost
# ---------------------------------------------------------------------------

PROCESSES_HEADERS = [
    "Batch_ID",
    "Process_Type",
    "Location",
    "Position_in_Plan",
    "Step_Index",
    "Step_Label",
    "Duration_Value",
    "Duration_Unit",
    "Rate_Value",
    "Rate_Unit",
    "Num_Samples_Covered",
    "Sample_IDs",
    "Notes",
]


def _build_processes_sheet(wb: Workbook, dm: LCCDataManager) -> None:
    ws = wb.create_sheet("Processes")
    _write_header(ws, PROCESSES_HEADERS)

    for row_index, process_row in enumerate(dm.process_rows, start=2):
        ws.cell(
            row=row_index, column=_col(PROCESSES_HEADERS, "Batch_ID"), value=process_row.batch_id
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Process_Type"),
            value=process_row.process_type,
        )
        ws.cell(
            row=row_index, column=_col(PROCESSES_HEADERS, "Location"), value=process_row.location
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Position_in_Plan"),
            value=process_row.position_in_plan,
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Step_Index"),
            value=process_row.step_index,
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Step_Label"),
            value=process_row.step_label,
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Duration_Value"),
            value=process_row.duration_value,
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Duration_Unit"),
            value=process_row.duration_unit,
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Rate_Value"),
            value=process_row.rate_value,
        )
        ws.cell(
            row=row_index, column=_col(PROCESSES_HEADERS, "Rate_Unit"), value=process_row.rate_unit
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Num_Samples_Covered"),
            value=process_row.num_samples_covered,
        )
        ws.cell(
            row=row_index,
            column=_col(PROCESSES_HEADERS, "Sample_IDs"),
            value=", ".join(process_row.sample_ids),
        )
        ws.cell(row=row_index, column=_col(PROCESSES_HEADERS, "Notes"), value="")

    _autosize_columns(ws, PROCESSES_HEADERS)


# ---------------------------------------------------------------------------
# Materials sheet
# ---------------------------------------------------------------------------

MATERIALS_HEADERS = [
    "Batch_ID",
    "Process_Type",
    "Material_Name",
    "Role",
    "CAS_Number",
    "Quantity_Value",
    "Quantity_Unit",
    "Molar_Mass_g_per_mol",
    "Quantity_Grams",
    "Price_per_Gram_Est",
    "Total_Price_Est",
    "Num_Samples_Covered",
    "Sample_IDs",
    "Cost_Low",
    "Cost_High",
    "Effective_Cost_Low",
    "Effective_Cost_Est",
    "Effective_Cost_High",
    "Cost_Per_Sample_Est",
    "Verified",
    "Notes",
]


def _build_materials_sheet(
    wb: Workbook, dm: LCCDataManager, cost_reference: CostReference | None
) -> None:
    ws = wb.create_sheet("Materials")
    _write_header(ws, MATERIALS_HEADERS)

    qty_grams_col = _letter(MATERIALS_HEADERS, "Quantity_Grams")
    price_col = _letter(MATERIALS_HEADERS, "Price_per_Gram_Est")
    total_price_col = _letter(MATERIALS_HEADERS, "Total_Price_Est")
    num_samples_col = _letter(MATERIALS_HEADERS, "Num_Samples_Covered")
    cost_low_col = _letter(MATERIALS_HEADERS, "Cost_Low")
    cost_high_col = _letter(MATERIALS_HEADERS, "Cost_High")
    eff_est_col = _letter(MATERIALS_HEADERS, "Effective_Cost_Est")
    verified_col = _letter(MATERIALS_HEADERS, "Verified")

    for row_index, material_row in enumerate(dm.material_rows, start=2):
        ref = (
            cost_reference.material_prices.get(material_row.material_name)
            if cost_reference
            else None
        )
        ref_price = ref.get("price_per_gram_est") if ref else None
        price_per_gram = ref_price if ref_price is not None else material_row.price_per_gram_est
        cas_number = (ref.get("cas_number") if ref else None) or material_row.cas_number

        ws.cell(
            row=row_index, column=_col(MATERIALS_HEADERS, "Batch_ID"), value=material_row.batch_id
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Process_Type"),
            value=material_row.process_type,
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Material_Name"),
            value=material_row.material_name,
        )
        ws.cell(row=row_index, column=_col(MATERIALS_HEADERS, "Role"), value=material_row.role)
        ws.cell(row=row_index, column=_col(MATERIALS_HEADERS, "CAS_Number"), value=cas_number)
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Quantity_Value"),
            value=material_row.quantity_value,
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Quantity_Unit"),
            value=material_row.quantity_unit,
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Molar_Mass_g_per_mol"),
            value=material_row.molar_mass_g_per_mol,
        )
        # Quantity_Grams left blank - see Guide sheet.
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Price_per_Gram_Est"),
            value=price_per_gram,
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Total_Price_Est"),
            value=(
                f'=IF(OR({qty_grams_col}{row_index}="",{price_col}{row_index}=""),"",'
                f"{qty_grams_col}{row_index}*{price_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Num_Samples_Covered"),
            value=material_row.num_samples_covered,
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Sample_IDs"),
            value=", ".join(material_row.sample_ids),
        )
        # Cost_Low/Cost_High left blank for manual entry.
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Effective_Cost_Low"),
            value=(
                f'=IF({cost_low_col}{row_index}="",{total_price_col}{row_index},'
                f"{cost_low_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Effective_Cost_Est"),
            value=f"={total_price_col}{row_index}",
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Effective_Cost_High"),
            value=(
                f'=IF({cost_high_col}{row_index}="",{total_price_col}{row_index},'
                f"{cost_high_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Cost_Per_Sample_Est"),
            value=(
                f'=IF({num_samples_col}{row_index}=0,"",'
                f"{eff_est_col}{row_index}/{num_samples_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(MATERIALS_HEADERS, "Verified"),
            value=bool(ref["verified"]) if ref else False,
        )
        if ref and ref.get("notes"):
            notes = ref["notes"]
        elif price_per_gram is not None:
            notes = CHEMICAL_REFERENCE_NOTE
        else:
            notes = ""
        ws.cell(row=row_index, column=_col(MATERIALS_HEADERS, "Notes"), value=notes)

    _add_bool_dropdown(ws, verified_col)
    _highlight_verified_column(ws, verified_col)
    _autosize_columns(ws, MATERIALS_HEADERS)


# ---------------------------------------------------------------------------
# Labor sheet - manual role-based entry (see Guide sheet for why)
# ---------------------------------------------------------------------------

LABOR_HEADERS = [
    "Batch_ID",
    "Role",
    "Hours",
    "Hourly_Rate_Est",
    "Cost_Est",
    "Num_Samples",
    "Cost_Low",
    "Cost_High",
    "Effective_Cost_Low",
    "Effective_Cost_Est",
    "Effective_Cost_High",
    "Cost_Per_Sample_Est",
    "Verified",
    "Notes",
]


def _build_labor_sheet(
    wb: Workbook, dm: LCCDataManager, cost_reference: CostReference | None
) -> None:
    ws = wb.create_sheet("Labor")
    _write_header(ws, LABOR_HEADERS)

    hours_col = _letter(LABOR_HEADERS, "Hours")
    rate_col = _letter(LABOR_HEADERS, "Hourly_Rate_Est")
    cost_est_col = _letter(LABOR_HEADERS, "Cost_Est")
    num_samples_col = _letter(LABOR_HEADERS, "Num_Samples")
    cost_low_col = _letter(LABOR_HEADERS, "Cost_Low")
    cost_high_col = _letter(LABOR_HEADERS, "Cost_High")
    eff_est_col = _letter(LABOR_HEADERS, "Effective_Cost_Est")
    role_col = _letter(LABOR_HEADERS, "Role")
    verified_col = _letter(LABOR_HEADERS, "Verified")

    row_index = 2
    for batch_id, num_samples in sorted(dm.batch_sample_counts.items()):
        for role in LABOR_ROLES:
            ref = cost_reference.labor_rates.get(role) if cost_reference else None
            ws.cell(row=row_index, column=_col(LABOR_HEADERS, "Batch_ID"), value=batch_id)
            ws.cell(row=row_index, column=_col(LABOR_HEADERS, "Role"), value=role)
            # Hours left blank for manual entry per batch.
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Hourly_Rate_Est"),
                value=ref["hourly_rate_est"] if ref else None,
            )
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Cost_Est"),
                value=(
                    f'=IF(OR({hours_col}{row_index}="",{rate_col}{row_index}=""),"",'
                    f"{hours_col}{row_index}*{rate_col}{row_index})"
                ),
            )
            ws.cell(row=row_index, column=_col(LABOR_HEADERS, "Num_Samples"), value=num_samples)
            # Cost_Low/Cost_High left blank for manual entry.
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Effective_Cost_Low"),
                value=(
                    f'=IF({cost_low_col}{row_index}="",{cost_est_col}{row_index},'
                    f"{cost_low_col}{row_index})"
                ),
            )
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Effective_Cost_Est"),
                value=f"={cost_est_col}{row_index}",
            )
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Effective_Cost_High"),
                value=(
                    f'=IF({cost_high_col}{row_index}="",{cost_est_col}{row_index},'
                    f"{cost_high_col}{row_index})"
                ),
            )
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Cost_Per_Sample_Est"),
                value=(
                    f'=IF({num_samples_col}{row_index}=0,"",'
                    f"{eff_est_col}{row_index}/{num_samples_col}{row_index})"
                ),
            )
            ws.cell(
                row=row_index,
                column=_col(LABOR_HEADERS, "Verified"),
                value=bool(ref["verified"]) if ref else False,
            )
            row_index += 1

    _add_list_dropdown(ws, role_col, LABOR_ROLES)
    _add_bool_dropdown(ws, verified_col)
    _highlight_verified_column(ws, verified_col)
    _autosize_columns(ws, LABOR_HEADERS)


# ---------------------------------------------------------------------------
# Capital / Overhead / Disposal sheet (fully manual entry - no NOMAD source)
# ---------------------------------------------------------------------------

CAPITAL_HEADERS = [
    "Batch_ID",
    "Item",
    "Location",
    "Cost_Category",
    "Allocation_Method",
    "Cost_Low",
    "Cost_Est",
    "Cost_High",
    "Num_Samples_Covered",
    "Effective_Cost_Low",
    "Effective_Cost_Est",
    "Effective_Cost_High",
    "Cost_Per_Sample_Est",
    "Verified",
    "Notes",
]


def _seed_capital_overhead_disposal_rows(dm: LCCDataManager) -> list[tuple]:
    """Template rows so colleagues have somewhere to start, scoped to what
    this batch selection actually used: one cleanroom rent placeholder per
    batch, one equipment-depreciation placeholder per distinct real NOMAD
    location seen (raw text, not normalized against KNOWN_LOCATIONS - see
    Guide sheet), one disposal placeholder per distinct material seen. No
    costs are fabricated - Cost_Low/Est/High stay blank.

    Each row: (batch_id, item, location, category, allocation_method, num_samples).
    """
    rows: list[tuple] = []
    for batch_id, num_samples in dm.batch_sample_counts.items():
        rows.append((batch_id, "Cleanroom / lab rent", "", "Overhead", "per hour", num_samples))

    locations_by_batch: dict[str, set[str]] = {}
    for row in dm.process_rows:
        if row.location:
            locations_by_batch.setdefault(row.batch_id, set()).add(row.location)
    for batch_id, locations in locations_by_batch.items():
        for location in sorted(locations):
            rows.append(
                (
                    batch_id,
                    f"Equipment depreciation - {location}",
                    location,
                    "Capital",
                    "per batch",
                    1,
                )
            )

    materials_by_batch: dict[str, set[str]] = {}
    for row in dm.material_rows:
        materials_by_batch.setdefault(row.batch_id, set()).add(row.material_name)
    for batch_id, material_names in materials_by_batch.items():
        for material_name in sorted(material_names):
            rows.append(
                (batch_id, f"Disposal - {material_name}", "", "End-of-Life", "per batch", 1)
            )

    return rows


def _build_capital_overhead_disposal_sheet(
    wb: Workbook, dm: LCCDataManager, cost_reference: CostReference | None
) -> None:
    ws = wb.create_sheet("Capital_Overhead_Disposal")
    _write_header(ws, CAPITAL_HEADERS)

    cost_low_col = _letter(CAPITAL_HEADERS, "Cost_Low")
    cost_est_col = _letter(CAPITAL_HEADERS, "Cost_Est")
    cost_high_col = _letter(CAPITAL_HEADERS, "Cost_High")
    num_samples_col = _letter(CAPITAL_HEADERS, "Num_Samples_Covered")
    eff_est_col = _letter(CAPITAL_HEADERS, "Effective_Cost_Est")
    category_col = _letter(CAPITAL_HEADERS, "Cost_Category")
    allocation_col = _letter(CAPITAL_HEADERS, "Allocation_Method")
    verified_col = _letter(CAPITAL_HEADERS, "Verified")

    for row_index, (batch_id, item, location, category, allocation, num_samples) in enumerate(
        _seed_capital_overhead_disposal_rows(dm), start=2
    ):
        ref = cost_reference.overhead_costs.get(item) if cost_reference else None

        ws.cell(row=row_index, column=_col(CAPITAL_HEADERS, "Batch_ID"), value=batch_id)
        ws.cell(row=row_index, column=_col(CAPITAL_HEADERS, "Item"), value=item)
        ws.cell(row=row_index, column=_col(CAPITAL_HEADERS, "Location"), value=location)
        ws.cell(row=row_index, column=_col(CAPITAL_HEADERS, "Cost_Category"), value=category)
        ws.cell(row=row_index, column=_col(CAPITAL_HEADERS, "Allocation_Method"), value=allocation)
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Cost_Low"),
            value=ref["cost_low"] if ref else None,
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Cost_Est"),
            value=ref["cost_est"] if ref else None,
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Cost_High"),
            value=ref["cost_high"] if ref else None,
        )
        ws.cell(
            row=row_index, column=_col(CAPITAL_HEADERS, "Num_Samples_Covered"), value=num_samples
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Effective_Cost_Low"),
            value=(
                f'=IF({cost_low_col}{row_index}="",{cost_est_col}{row_index},'
                f"{cost_low_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Effective_Cost_Est"),
            value=f"={cost_est_col}{row_index}",
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Effective_Cost_High"),
            value=(
                f'=IF({cost_high_col}{row_index}="",{cost_est_col}{row_index},'
                f"{cost_high_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Cost_Per_Sample_Est"),
            value=(
                f'=IF({num_samples_col}{row_index}=0,"",'
                f"{eff_est_col}{row_index}/{num_samples_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(CAPITAL_HEADERS, "Verified"),
            value=bool(ref["verified"]) if ref else False,
        )
        ws.cell(
            row=row_index, column=_col(CAPITAL_HEADERS, "Notes"), value=ref["notes"] if ref else ""
        )

    _add_list_dropdown(ws, category_col, COST_CATEGORY_OPTIONS)
    _add_list_dropdown(ws, allocation_col, ALLOCATION_METHOD_OPTIONS)
    _add_list_dropdown(ws, _letter(CAPITAL_HEADERS, "Location"), [*KNOWN_LOCATIONS, ""])
    _add_bool_dropdown(ws, verified_col)
    _highlight_verified_column(ws, verified_col)
    _autosize_columns(ws, CAPITAL_HEADERS)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

SUMMARY_HEADERS = [
    "Batch_ID",
    "Num_Samples",
    "Total_Material_Cost_Est",
    "Total_Labor_Cost_Est",
    "Total_Capital_Overhead_Disposal_Cost_Est",
    "Grand_Total_Low",
    "Grand_Total_Est",
    "Grand_Total_High",
    "Grand_Total_Per_Sample_Est",
    "Unverified_Line_Items",
    "Data_Quality",
]

_UNVERIFIED_FLAG = "⚠ Contains unverified values"
_VERIFIED_FLAG = "✓ Verified"


def _build_summary_sheet(wb: Workbook, dm: LCCDataManager) -> None:
    ws = wb.create_sheet("Summary")
    _write_header(ws, SUMMARY_HEADERS)

    mat_batch = _letter(MATERIALS_HEADERS, "Batch_ID")
    mat_eff_low = _letter(MATERIALS_HEADERS, "Effective_Cost_Low")
    mat_eff_est = _letter(MATERIALS_HEADERS, "Effective_Cost_Est")
    mat_eff_high = _letter(MATERIALS_HEADERS, "Effective_Cost_High")
    mat_verified = _letter(MATERIALS_HEADERS, "Verified")

    labor_batch = _letter(LABOR_HEADERS, "Batch_ID")
    labor_eff_low = _letter(LABOR_HEADERS, "Effective_Cost_Low")
    labor_eff_est = _letter(LABOR_HEADERS, "Effective_Cost_Est")
    labor_eff_high = _letter(LABOR_HEADERS, "Effective_Cost_High")
    labor_verified = _letter(LABOR_HEADERS, "Verified")

    cap_batch = _letter(CAPITAL_HEADERS, "Batch_ID")
    cap_eff_low = _letter(CAPITAL_HEADERS, "Effective_Cost_Low")
    cap_eff_est = _letter(CAPITAL_HEADERS, "Effective_Cost_Est")
    cap_eff_high = _letter(CAPITAL_HEADERS, "Effective_Cost_High")
    cap_verified = _letter(CAPITAL_HEADERS, "Verified")

    batch_col = _letter(SUMMARY_HEADERS, "Batch_ID")

    for row_index, (batch_id, num_samples) in enumerate(
        sorted(dm.batch_sample_counts.items()), start=2
    ):
        batch_ref = f"{batch_col}{row_index}"
        ws.cell(row=row_index, column=_col(SUMMARY_HEADERS, "Batch_ID"), value=batch_id)
        ws.cell(row=row_index, column=_col(SUMMARY_HEADERS, "Num_Samples"), value=num_samples)
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Total_Material_Cost_Est"),
            value=f"=SUMIFS(Materials!{mat_eff_est}:{mat_eff_est},Materials!{mat_batch}:{mat_batch},{batch_ref})",
        )
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Total_Labor_Cost_Est"),
            value=f"=SUMIFS(Labor!{labor_eff_est}:{labor_eff_est},Labor!{labor_batch}:{labor_batch},{batch_ref})",
        )
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Total_Capital_Overhead_Disposal_Cost_Est"),
            value=(
                f"=SUMIFS(Capital_Overhead_Disposal!{cap_eff_est}:{cap_eff_est},"
                f"Capital_Overhead_Disposal!{cap_batch}:{cap_batch},{batch_ref})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Grand_Total_Low"),
            value=(
                f"=SUMIFS(Materials!{mat_eff_low}:{mat_eff_low},Materials!{mat_batch}:{mat_batch},{batch_ref})"
                f"+SUMIFS(Labor!{labor_eff_low}:{labor_eff_low},Labor!{labor_batch}:{labor_batch},{batch_ref})"
                f"+SUMIFS(Capital_Overhead_Disposal!{cap_eff_low}:{cap_eff_low},"
                f"Capital_Overhead_Disposal!{cap_batch}:{cap_batch},{batch_ref})"
            ),
        )
        total_col = _letter(SUMMARY_HEADERS, "Total_Material_Cost_Est")
        labor_total_col = _letter(SUMMARY_HEADERS, "Total_Labor_Cost_Est")
        cap_total_col = _letter(SUMMARY_HEADERS, "Total_Capital_Overhead_Disposal_Cost_Est")
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Grand_Total_Est"),
            value=(
                f"={total_col}{row_index}+{labor_total_col}{row_index}+{cap_total_col}{row_index}"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Grand_Total_High"),
            value=(
                f"=SUMIFS(Materials!{mat_eff_high}:{mat_eff_high},Materials!{mat_batch}:{mat_batch},{batch_ref})"
                f"+SUMIFS(Labor!{labor_eff_high}:{labor_eff_high},Labor!{labor_batch}:{labor_batch},{batch_ref})"
                f"+SUMIFS(Capital_Overhead_Disposal!{cap_eff_high}:{cap_eff_high},"
                f"Capital_Overhead_Disposal!{cap_batch}:{cap_batch},{batch_ref})"
            ),
        )
        num_samples_col = _letter(SUMMARY_HEADERS, "Num_Samples")
        grand_total_est_col = _letter(SUMMARY_HEADERS, "Grand_Total_Est")
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Grand_Total_Per_Sample_Est"),
            value=(
                f'=IF({num_samples_col}{row_index}=0,"",'
                f"{grand_total_est_col}{row_index}/{num_samples_col}{row_index})"
            ),
        )
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Unverified_Line_Items"),
            value=(
                f"=COUNTIFS(Materials!{mat_batch}:{mat_batch},{batch_ref},"
                f"Materials!{mat_verified}:{mat_verified},FALSE)"
                f"+COUNTIFS(Labor!{labor_batch}:{labor_batch},{batch_ref},"
                f"Labor!{labor_verified}:{labor_verified},FALSE)"
                f"+COUNTIFS(Capital_Overhead_Disposal!{cap_batch}:{cap_batch},{batch_ref},"
                f"Capital_Overhead_Disposal!{cap_verified}:{cap_verified},FALSE)"
            ),
        )
        unverified_col = _letter(SUMMARY_HEADERS, "Unverified_Line_Items")
        ws.cell(
            row=row_index,
            column=_col(SUMMARY_HEADERS, "Data_Quality"),
            value=(f'=IF({unverified_col}{row_index}>0,"{_UNVERIFIED_FLAG}","{_VERIFIED_FLAG}")'),
        )

    data_quality_col = _letter(SUMMARY_HEADERS, "Data_Quality")
    ws.conditional_formatting.add(
        f"{data_quality_col}2:{data_quality_col}{MAX_DATA_ROW}",
        CellIsRule(operator="equal", formula=[f'"{_UNVERIFIED_FLAG}"'], fill=WARNING_FILL),
    )
    _autosize_columns(ws, SUMMARY_HEADERS)


# ---------------------------------------------------------------------------
# Cost reference template - the admin-maintained single source of truth
# ---------------------------------------------------------------------------
# A separate, lean workbook shape from the per-batch export above: no
# Batch_ID/Sample_IDs/per-instance columns, just the reusable lookup a
# material name / role / overhead item carries forward.
# data_manager.parse_cost_reference_workbook reads this by header name, so
# it works equally well on this template or on a full per-batch export.

_REFERENCE_GUIDE_TEXT = [
    ("LCC Cost Reference - Guide (for admins)", True),
    ("", False),
    (
        "This is the single source of truth for cost figures. Every user's "
        "LCC Calculator session reads this same file automatically when "
        "exporting a batch's costing workbook - there is nothing for them "
        "to upload or configure.",
        False,
    ),
    ("", False),
    (
        "To update a cost: edit the matching row's Cost_Est (and Low/High "
        "if you have a range) on the relevant sheet below, set Verified to "
        "TRUE, and save this file in place. The change is live for every "
        "user on their next export.",
        False,
    ),
    ("", False),
    (
        "To add a new material/role/item: add a new row on the relevant "
        "sheet with a name that exactly matches what the app extracts from "
        "NOMAD (check an exported batch workbook's Materials/Capital_"
        "Overhead_Disposal sheets for the exact names in use) - matching is "
        "by exact text, so a typo means the row is silently ignored rather "
        "than applied. Labor is the exception: Role is always one of the "
        "4 fixed tiers, no typos possible.",
        False,
    ),
    ("", False),
    (
        "Materials sheet - Price and Grams_on_Bottle: enter what you paid "
        "for a bottle/container (Price) and how many grams it holds "
        "(Grams_on_Bottle) - Price_per_Gram_Est is computed automatically. "
        "Pre-seeded with rough starting estimates for common perovskite-fab "
        "chemicals (Verified=FALSE) - replace with your actual supplier "
        "prices and mark Verified=TRUE as you confirm each one.",
        False,
    ),
    ("", False),
    (
        "Capital_Overhead_Disposal sheet - Location: pre-seeded with one "
        "equipment-depreciation placeholder per known glovebox/tool. See "
        "the per-batch export's Guide sheet for why raw NOMAD location text "
        "doesn't always match these names automatically.",
        False,
    ),
]

_REFERENCE_MATERIALS_HEADERS = [
    "Material_Name",
    "CAS_Number",
    "Price",
    "Grams_on_Bottle",
    "Price_per_Gram_Est",
    "Verified",
    "Notes",
]
_REFERENCE_PROCESSES_HEADERS = ["Process_Type", "Step_Label", "Notes"]
_REFERENCE_LABOR_HEADERS = ["Role", "Hourly_Rate_Est", "Verified"]
_REFERENCE_CAPITAL_HEADERS = [
    "Item",
    "Location",
    "Category",
    "Allocation_Method",
    "Cost_Low",
    "Cost_Est",
    "Cost_High",
    "Verified",
    "Notes",
]


def _write_reference_guide_sheet(wb: Workbook, generated_from_note: str) -> None:
    ws = wb.active
    ws.title = "Guide"
    lines = [
        *_REFERENCE_GUIDE_TEXT[:2],
        (f"Generated from: {generated_from_note}.", False),
        ("", False),
        *_REFERENCE_GUIDE_TEXT[2:],
    ]
    for row_index, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=row_index, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13 if row_index == 1 else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def _write_reference_materials_sheet(
    wb: Workbook, entries: list[tuple[str, str | None, float | None, str]]
) -> None:
    """entries: (material_name, cas_number, price_per_gram_est, notes). Price
    is seeded equal to price_per_gram_est with Grams_on_Bottle=1 (i.e. "price
    per gram" expressed as a 1-gram bottle) so the formula-computed
    Price_per_Gram_Est matches today's static estimate exactly."""
    ws = wb.create_sheet("Materials")
    _write_header(ws, _REFERENCE_MATERIALS_HEADERS)
    price_col = _letter(_REFERENCE_MATERIALS_HEADERS, "Price")
    grams_col = _letter(_REFERENCE_MATERIALS_HEADERS, "Grams_on_Bottle")
    for row_index, (name, cas_number, price_per_gram, notes) in enumerate(entries, start=2):
        ws.cell(
            row=row_index, column=_col(_REFERENCE_MATERIALS_HEADERS, "Material_Name"), value=name
        )
        ws.cell(
            row=row_index, column=_col(_REFERENCE_MATERIALS_HEADERS, "CAS_Number"), value=cas_number
        )
        ws.cell(
            row=row_index, column=_col(_REFERENCE_MATERIALS_HEADERS, "Price"), value=price_per_gram
        )
        ws.cell(
            row=row_index,
            column=_col(_REFERENCE_MATERIALS_HEADERS, "Grams_on_Bottle"),
            value=1 if price_per_gram is not None else None,
        )
        ws.cell(
            row=row_index,
            column=_col(_REFERENCE_MATERIALS_HEADERS, "Price_per_Gram_Est"),
            value=(
                f'=IF(OR({price_col}{row_index}="",{grams_col}{row_index}=""),"",'
                f"{price_col}{row_index}/{grams_col}{row_index})"
            ),
        )
        ws.cell(row=row_index, column=_col(_REFERENCE_MATERIALS_HEADERS, "Verified"), value=False)
        ws.cell(row=row_index, column=_col(_REFERENCE_MATERIALS_HEADERS, "Notes"), value=notes)
    verified_col = _letter(_REFERENCE_MATERIALS_HEADERS, "Verified")
    _add_bool_dropdown(ws, verified_col)
    _highlight_verified_column(ws, verified_col)
    _autosize_columns(ws, _REFERENCE_MATERIALS_HEADERS)


def _write_reference_processes_sheet(wb: Workbook, entries: list[tuple[str, str]]) -> None:
    """entries: (process_type, step_label). Informational catalog only - no
    cost columns, see Guide sheet for why."""
    ws = wb.create_sheet("Processes")
    _write_header(ws, _REFERENCE_PROCESSES_HEADERS)
    for row_index, (process_type, step_label) in enumerate(entries, start=2):
        ws.cell(
            row=row_index,
            column=_col(_REFERENCE_PROCESSES_HEADERS, "Process_Type"),
            value=process_type,
        )
        ws.cell(
            row=row_index, column=_col(_REFERENCE_PROCESSES_HEADERS, "Step_Label"), value=step_label
        )
    _autosize_columns(ws, _REFERENCE_PROCESSES_HEADERS)


def _write_reference_labor_sheet(wb: Workbook) -> None:
    """Always the 4 fixed role tiers - see data_manager.LABOR_ROLES."""
    ws = wb.create_sheet("Labor")
    _write_header(ws, _REFERENCE_LABOR_HEADERS)
    for row_index, role in enumerate(LABOR_ROLES, start=2):
        ws.cell(row=row_index, column=_col(_REFERENCE_LABOR_HEADERS, "Role"), value=role)
        ws.cell(row=row_index, column=_col(_REFERENCE_LABOR_HEADERS, "Verified"), value=False)
    verified_col = _letter(_REFERENCE_LABOR_HEADERS, "Verified")
    _add_bool_dropdown(ws, verified_col)
    _highlight_verified_column(ws, verified_col)
    _autosize_columns(ws, _REFERENCE_LABOR_HEADERS)


def _write_reference_capital_sheet(wb: Workbook, entries: list[tuple[str, str, str, str]]) -> None:
    """entries: (item, location, category, allocation_method)."""
    ws = wb.create_sheet("Capital_Overhead_Disposal")
    _write_header(ws, _REFERENCE_CAPITAL_HEADERS)
    for row_index, (item, location, category, allocation) in enumerate(entries, start=2):
        ws.cell(row=row_index, column=_col(_REFERENCE_CAPITAL_HEADERS, "Item"), value=item)
        ws.cell(row=row_index, column=_col(_REFERENCE_CAPITAL_HEADERS, "Location"), value=location)
        ws.cell(row=row_index, column=_col(_REFERENCE_CAPITAL_HEADERS, "Category"), value=category)
        ws.cell(
            row=row_index,
            column=_col(_REFERENCE_CAPITAL_HEADERS, "Allocation_Method"),
            value=allocation,
        )
        ws.cell(row=row_index, column=_col(_REFERENCE_CAPITAL_HEADERS, "Verified"), value=False)
    verified_col = _letter(_REFERENCE_CAPITAL_HEADERS, "Verified")
    _add_list_dropdown(ws, _letter(_REFERENCE_CAPITAL_HEADERS, "Category"), COST_CATEGORY_OPTIONS)
    _add_list_dropdown(
        ws, _letter(_REFERENCE_CAPITAL_HEADERS, "Allocation_Method"), ALLOCATION_METHOD_OPTIONS
    )
    _add_list_dropdown(ws, _letter(_REFERENCE_CAPITAL_HEADERS, "Location"), [*KNOWN_LOCATIONS, ""])
    _add_bool_dropdown(ws, verified_col)
    _highlight_verified_column(ws, verified_col)
    _autosize_columns(ws, _REFERENCE_CAPITAL_HEADERS)


def _known_location_capital_entries() -> list[tuple[str, str, str, str]]:
    entries = [("Cleanroom / lab rent", "", "Overhead", "per hour")]
    entries.extend(
        (f"Equipment depreciation - {location}", location, "Capital", "per batch")
        for location in KNOWN_LOCATIONS
    )
    return entries


def build_cost_reference_template() -> Workbook:
    """A generic starting point for the admin-maintained master cost file
    (data_manager.DEFAULT_COST_REFERENCE_PATH) - Materials pre-seeded from
    the static CHEMICAL_REFERENCE table only (no NOMAD data involved),
    Capital_Overhead_Disposal pre-seeded from the fixed KNOWN_LOCATIONS
    list. Prefer build_cost_reference_template_from_data when real batch
    data is available - it guarantees exact-match material/process names
    instead of relying on manual typing.
    """
    wb = Workbook()
    _write_reference_guide_sheet(
        wb, "the static chemical reference list and known locations only (no NOMAD data scanned)"
    )
    materials_entries = [
        (name, info["cas_number"], info["price_per_gram_est"], CHEMICAL_REFERENCE_NOTE)
        for name, info in sorted(CHEMICAL_REFERENCE.items())
    ]
    _write_reference_materials_sheet(wb, materials_entries)
    _write_reference_processes_sheet(wb, [])
    _write_reference_labor_sheet(wb)
    _write_reference_capital_sheet(wb, _known_location_capital_entries())
    return wb


def build_cost_reference_template_from_data(
    dm: LCCDataManager, extra_schema_types: list[str] | None = None
) -> Workbook:
    """Same shape as build_cost_reference_template, but Materials/Processes
    rows are seeded from every distinct Material_Name / (Process_Type,
    Step_Label) actually found in dm's currently loaded batches - guarantees
    exact-match names against what the app extracts from NOMAD, so an admin
    never has to hand-type (and risk mistyping) a name. Falls back to the
    static CHEMICAL_REFERENCE list for any common chemical not yet
    encountered in the loaded batches. extra_schema_types (e.g. from
    data_manager.discover_entry_types) adds measurement/other schemas that
    have no step-level detail to extract, listed with a blank Step_Label.

    Labor and Capital_Overhead_Disposal locations are NOT derived from data -
    they're always the fixed role tiers / known locations, since those are
    stable lab facts independent of which batches happen to be selected.
    Intended for an admin to run over a broad/complete batch selection, then
    save the result as data_manager.DEFAULT_COST_REFERENCE_PATH.
    """
    wb = Workbook()
    batch_count = len(dm.batch_sample_counts)
    _write_reference_guide_sheet(
        wb,
        f"{batch_count} selected batch(es) - re-run with a broader batch "
        "selection (ideally all batches) for fuller coverage",
    )

    seen_materials = {row.material_name: row.cas_number for row in dm.material_rows}
    all_material_names = sorted(set(seen_materials) | set(CHEMICAL_REFERENCE.keys()))
    materials_entries = [
        (
            name,
            seen_materials.get(name)
            or (CHEMICAL_REFERENCE[name]["cas_number"] if name in CHEMICAL_REFERENCE else None),
            CHEMICAL_REFERENCE[name]["price_per_gram_est"] if name in CHEMICAL_REFERENCE else None,
            CHEMICAL_REFERENCE_NOTE if name in CHEMICAL_REFERENCE else "",
        )
        for name in all_material_names
    ]
    _write_reference_materials_sheet(wb, materials_entries)

    process_entries = sorted({(row.process_type, row.step_label) for row in dm.process_rows})
    known_process_types = {process_type for process_type, _ in process_entries}
    for schema_type in sorted(set(extra_schema_types or []) - known_process_types):
        process_entries.append((schema_type, ""))
    _write_reference_processes_sheet(wb, process_entries)

    _write_reference_labor_sheet(wb)
    _write_reference_capital_sheet(wb, _known_location_capital_entries())

    return wb


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def build_workbook(dm: LCCDataManager, cost_reference: CostReference | None = None) -> Workbook:
    """cost_reference, when given (see data_manager.load_default_cost_reference),
    carries forward matching costs from the shared admin-maintained master cost
    file instead of leaving every cell blank/unverified again."""
    wb = Workbook()
    _build_guide_sheet(wb)
    _build_processes_sheet(wb, dm)
    _build_materials_sheet(wb, dm, cost_reference)
    _build_labor_sheet(wb, dm, cost_reference)
    _build_capital_overhead_disposal_sheet(wb, dm, cost_reference)
    _build_summary_sheet(wb, dm)
    logger.info("Built LCC workbook: %s", wb.sheetnames)
    return wb

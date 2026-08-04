"""
data_manager.py
----------------
Pure Python / Pydantic layer for the LCC (Life Cycle Costing) Calculator.
No widget imports. Fetches processes/materials for selected batches from
NOMAD, extracts cost-relevant rows, and provides the cost-range/per-sample
math used by excel_export.py.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import requests
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Domain models
# ---------------------------------------------------------------------------

# Fixed labor role tiers - NOMAD has no field tracking who (or what role)
# ran a process (confirmed against live data: `operator` is populated on
# roughly half of process entries and is only ever a free-text name, never
# a role), so labor cost cannot be auto-extracted per process instance.
# Instead a batch's labor cost is entered manually per role (hours worked),
# with the hourly rate coming from the shared cost reference file.
LABOR_ROLES = ["PhD Researcher", "Postdoc", "Engineer", "HiWi"]

# Canonical glovebox/tool names, as given by the lab (not derived from
# NOMAD's `location` field, which is free text typed by whoever ran the
# process - confirmed against live data to have many inconsistent variants
# of the same physical box, e.g. "HySpinBox"/"HySpin GB"/"Hyspin"/"hyspin").
# Used only as the Capital_Overhead_Disposal Location dropdown's option
# list - not matched automatically against the raw NOMAD text.
KNOWN_LOCATIONS = [
    "HyWeighBox",
    "HySolveBox",
    "HySpinBox",
    "HyPeroSpinBox",
    "HyALDBox",
    "HyVapBox",
    "HyPeroVapBox",
    "HyMessBox",
    "HySnPbBox",
    "ProtoSolveBox",
    "ProtoVapBox",
    "TinSpinBox",
    "TinVapBox",
    "TinHallBox",
    "InkCoatBox",
    "InkRollBox",
    "InkVapBox",
    "InkPatternBox",
]


class ProcessStepRow(BaseModel):
    batch_id: str
    process_type: str
    location: str = ""  # raw NOMAD text, informational only - see KNOWN_LOCATIONS note above
    position_in_plan: float | None = None
    step_index: int = 0
    step_label: str = ""
    duration_value: float | None = None
    duration_unit: str = ""
    rate_value: float | None = None
    rate_unit: str = ""
    num_samples_covered: int = 1
    sample_ids: list[str] = Field(default_factory=list)


class MaterialRow(BaseModel):
    batch_id: str
    process_type: str
    material_name: str
    role: str
    quantity_value: float | None = None
    quantity_unit: str = ""
    cas_number: str | None = None
    molar_mass_g_per_mol: float | None = None
    price_per_gram_est: float | None = None
    num_samples_covered: int = 1
    sample_ids: list[str] = Field(default_factory=list)


# ---------------------------------------------------------------------------
# Static chemical reference table
# ---------------------------------------------------------------------------
# No price data exists anywhere in this repo (confirmed by search), and live
# NOMAD data almost never has cas_number/molar_mass populated on chemical
# references either (checked against real entries: the fields exist in the
# schema but come back empty). CAS numbers below are real, well-established
# chemistry (safe to trust); price_per_gram_est is a rough general-knowledge
# estimate, NOT a live supplier lookup - every row using it is written to
# Excel with Verified=FALSE and a note asking colleagues to confirm.

# (name, cas_number, molar_mass_g_per_mol, price_per_gram_est)
_CHEMICAL_REFERENCE_ROWS: list[tuple[str, str, float, float]] = [
    ("PbI2", "10101-63-0", 461.01, 3.0),
    ("PbBr2", "10031-22-8", 367.01, 3.0),
    ("MAI", "14965-49-2", 158.97, 15.0),
    ("MABr", "18617-60-8", 111.97, 15.0),
    ("FAI", "879643-71-7", 186.99, 20.0),
    ("FABr", "1197196-79-2", 153.99, 20.0),
    ("MACl", "593-51-1", 67.52, 5.0),
    ("CsI", "7789-17-5", 259.81, 2.0),
    ("Cs2CO3", "534-17-8", 325.82, 3.0),
    ("Spiro-OMeTAD", "741909-64-2", 1225.44, 400.0),
    ("Li-TFSI", "90076-65-6", 287.09, 10.0),
    ("4-tert-Butylpyridine", "3978-81-2", 135.21, 1.0),
    ("DMF", "68-12-2", 73.09, 0.1),
    ("DMSO", "67-68-5", 78.13, 0.1),
    ("Chlorobenzene", "108-90-7", 112.56, 0.1),
    ("GBL", "96-48-0", 86.09, 0.2),
    ("IPA", "67-63-0", 60.10, 0.05),
    ("Toluene", "108-88-3", 92.14, 0.05),
    ("PCBM", "160848-22-6", 910.88, 100.0),
    ("C60", "99685-96-8", 720.66, 50.0),
]

CHEMICAL_REFERENCE: dict[str, dict] = {
    name: {"cas_number": cas, "molar_mass_g_per_mol": molar_mass, "price_per_gram_est": price}
    for name, cas, molar_mass, price in _CHEMICAL_REFERENCE_ROWS
}

CHEMICAL_REFERENCE_NOTE = (
    "Rough estimate based on general market knowledge, not a live supplier "
    "lookup - please confirm against your actual supplier/contract price."
)

# ---------------------------------------------------------------------------
# Generic flattening (adapted from Global_analyzer's _flatten_process_entry;
# reimplemented locally since apps don't import each other's app-local
# modules - see plan notes)
# ---------------------------------------------------------------------------


def flatten_entry(entry: dict, prefix: str = "") -> dict:
    """Recursively flatten a nested dict/list into dot-path keys.

    Single-item lists are unwrapped in place; multi-item lists get numeric
    ``.1``, ``.2``, ... suffixes so repeated sub-structures (e.g. recipe
    steps) become distinguishable keys instead of colliding.
    """
    out: dict = {}
    for key, value in entry.items():
        path = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            out.update(flatten_entry(value, path))
        elif isinstance(value, list):
            if not value:
                continue
            if len(value) == 1:
                item = value[0]
                if isinstance(item, dict):
                    out.update(flatten_entry(item, path))
                else:
                    out[path] = item
            else:
                for index, item in enumerate(value, start=1):
                    if isinstance(item, dict):
                        out.update(flatten_entry(item, f"{path}.{index}"))
                    else:
                        out[f"{path}.{index}"] = item
        elif isinstance(value, (str, int, float, bool)) or value is None:
            out[path] = value
    return out


_STEP_INDEX_RE = re.compile(r"^(?P<prefix>.+)\.(?P<index>\d+)\.(?P<field>.+)$")

_TIME_KEY_HINTS = ("time", "duration")
_RATE_KEY_HINTS = ("rate", "speed", "frequency")
_STEP_LIST_NAME_HINTS = ("step",)
_STEP_LIST_ALLOWLIST = {"cleaning", "cleaning_uv", "organic_evaporation"}

_SKIP_TOP_LEVEL_KEYS = {
    "samples",
    "m_def",
    "positon_in_experimental_plan",
    "datetime",
    "name",
    "description",
    "location",
    "operator",
    "lab_id",
}


def _group_into_steps(flat: dict) -> dict[tuple[str, int], dict]:
    """Group flattened dot-path keys by their list-index segment.

    Keys with no numeric list-index segment (e.g. a single-step process)
    fall under the sentinel key ("", 0).
    """
    steps: dict[tuple[str, int], dict] = {}
    for key, value in flat.items():
        match = _STEP_INDEX_RE.match(key)
        if match:
            step_key = (match.group("prefix"), int(match.group("index")))
            step_field = match.group("field")
        else:
            step_key = ("", 0)
            step_field = key
        steps.setdefault(step_key, {})[step_field] = value
    return steps


def _is_step_list_prefix(prefix: str) -> bool:
    if not prefix:
        return False
    last_segment = prefix.rsplit(".", 1)[-1].lower()
    return (
        any(hint in last_segment for hint in _STEP_LIST_NAME_HINTS)
        or last_segment in _STEP_LIST_ALLOWLIST
    )


def _first_numeric_match(fields: dict, hints: tuple[str, ...]) -> tuple[str, float] | None:
    for key, value in fields.items():
        if isinstance(value, bool):
            continue
        if isinstance(value, (int, float)) and any(hint in key.lower() for hint in hints):
            return key, float(value)
    return None


def extract_process_step_rows(
    process_data: dict,
    process_type: str,
    batch_id: str,
    num_samples_covered: int,
    sample_ids: list[str],
) -> list[ProcessStepRow]:
    """One row for the process itself (step_index 0), plus one row per
    detected step-like sub-list (e.g. recipe_steps, cleaning) - not every
    multi-item list, since solute/solvent lists belong to Materials, not
    Processes. Duration/rate only - no cost columns here (process cost is
    captured via Materials + Labor + Capital_Overhead_Disposal instead).
    """
    relevant = {k: v for k, v in process_data.items() if k not in _SKIP_TOP_LEVEL_KEYS}
    flat = flatten_entry(relevant)
    grouped = _group_into_steps(flat)
    position = process_data.get("positon_in_experimental_plan")
    location = str(process_data.get("location") or "").strip()

    rows: list[ProcessStepRow] = []
    for (prefix, index), fields in sorted(grouped.items(), key=lambda item: item[0]):
        if prefix and not _is_step_list_prefix(prefix):
            continue
        duration = _first_numeric_match(fields, _TIME_KEY_HINTS)
        rate = _first_numeric_match(fields, _RATE_KEY_HINTS)
        rows.append(
            ProcessStepRow(
                batch_id=batch_id,
                process_type=process_type,
                location=location,
                position_in_plan=position,
                step_index=index,
                step_label=prefix or process_type,
                duration_value=duration[1] if duration else None,
                duration_unit="as provided by NOMAD" if duration else "",
                rate_value=rate[1] if rate else None,
                rate_unit="as provided by NOMAD" if rate else "",
                num_samples_covered=num_samples_covered,
                sample_ids=list(sample_ids),
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Material extraction (known NOMAD field shapes, see plan research notes)
# ---------------------------------------------------------------------------


def extract_material_rows(
    process_data: dict,
    process_type: str,
    batch_id: str,
    num_samples_covered: int,
    sample_ids: list[str],
) -> list[MaterialRow]:
    rows: list[MaterialRow] = []

    def add(
        material_name,
        role: str,
        quantity_value,
        quantity_unit: str,
        chemical: dict | None = None,
    ):
        if not material_name or not str(material_name).strip():
            return
        name = str(material_name).strip()
        static_ref = CHEMICAL_REFERENCE.get(name)

        # A real cas_number/molar_mass straight from NOMAD (when a chemist
        # actually ran the PubChem lookup for this reference) always wins
        # over the static guess table - it's authoritative, not an estimate.
        chemical = chemical or {}
        static_cas = static_ref["cas_number"] if static_ref else None
        static_molar_mass = static_ref["molar_mass_g_per_mol"] if static_ref else None
        cas_number = chemical.get("cas_number") or static_cas
        molar_mass = chemical.get("molar_mass") or static_molar_mass

        rows.append(
            MaterialRow(
                batch_id=batch_id,
                process_type=process_type,
                material_name=name,
                role=role,
                quantity_value=quantity_value,
                quantity_unit=quantity_unit,
                cas_number=cas_number,
                molar_mass_g_per_mol=molar_mass,
                price_per_gram_est=static_ref["price_per_gram_est"] if static_ref else None,
                num_samples_covered=num_samples_covered,
                sample_ids=list(sample_ids),
            )
        )

    layers = process_data.get("layer") or []
    if layers and isinstance(layers[0], dict):
        add(layers[0].get("layer_material_name"), "layer", None, "")

    solution_list = process_data.get("solution") or []
    if solution_list and isinstance(solution_list[0], dict):
        details = solution_list[0].get("solution_details") or {}
        for solute in details.get("solute") or []:
            if not isinstance(solute, dict):
                continue
            chemical = solute.get("chemical_2") or {}
            add(
                chemical.get("name"),
                "solute",
                solute.get("concentration_mol"),
                "mol (as provided)",
                chemical,
            )
        for solvent in details.get("solvent") or []:
            if not isinstance(solvent, dict):
                continue
            chemical = solvent.get("chemical_2") or {}
            add(
                chemical.get("name"),
                "solvent",
                solvent.get("chemical_volume"),
                "volume (as provided)",
                chemical,
            )

    for evap in process_data.get("organic_evaporation") or []:
        if not isinstance(evap, dict):
            continue
        chemical = evap.get("chemical_2") or {}
        add(
            chemical.get("name"),
            "evaporant",
            evap.get("thickness"),
            "thickness (as provided)",
            chemical,
        )

    substrate = process_data.get("substrate_data") or {}
    conducting = substrate.get("conducting_material")
    if isinstance(conducting, list) and conducting:
        add(", ".join(str(c) for c in conducting), "substrate", None, "")

    return rows


# ---------------------------------------------------------------------------
# NOMAD API access
# ---------------------------------------------------------------------------


def fetch_samples_per_batch(url: str, token: str, batch_ids: list[str]) -> dict[str, list[str]]:
    """Map each batch id to its own sample ids.

    hysprint_utils.api_calls.get_ids_in_batch returns one flat list across
    all requested batches, which loses the per-batch grouping this app
    needs (batch sample counts, per-sample cost allocation) - so batches
    are queried one at a time here instead. get_ids_in_batch asserts the
    query returns exactly one batch record; confirmed against live data
    that at least one real batch fails this (a stale/unusual batch entry),
    which would otherwise crash the whole scan over an unrelated batch -
    such a batch is skipped (logged), not fatal.
    """
    from hysprint_utils.api_calls import get_ids_in_batch

    result: dict[str, list[str]] = {}
    for batch_id in batch_ids:
        try:
            result[batch_id] = get_ids_in_batch(url, token, [batch_id])
        except AssertionError:
            logger.warning("Skipping batch %s - could not resolve its samples via NOMAD.", batch_id)
    return result


def fetch_process_entries(
    url: str, token: str, sample_ids: list[str], process_type: str = "baseclasses.BaseProcess"
) -> list[tuple[dict, str]]:
    """Fetch every process entry linked to sample_ids, keeping the full
    ``samples`` list on each entry (hysprint_utils.api_calls.get_processing_steps
    and get_all_eqe only keep samples[0] - not reusable here since knowing how
    many samples a single process instance actually covers is exactly what
    lets 'one evaporation run for 16 samples' cost 1x per batch and 1/16th
    per sample instead of a guess).

    Returns a list of (process_data, entry_type) tuples, sorted by
    positon_in_experimental_plan (NOMAD's field name, not a typo introduced
    here).
    """
    headers = {"Authorization": f"Bearer {token}"}

    entry_id_query = {
        "required": {"metadata": "*"},
        "owner": "visible",
        "query": {"results.eln.lab_ids:any": sample_ids},
        "pagination": {"page_size": 10000},
    }
    response = requests.post(f"{url}/entries/query", headers=headers, json=entry_id_query)
    response.raise_for_status()
    entry_ids = [entry["entry_id"] for entry in response.json()["data"]]

    process_query = {
        "required": {"data": "*", "metadata": "*"},
        "owner": "visible",
        "query": {
            "entry_references.target_entry_id:any": entry_ids,
            "section_defs.definition_qualified_name": process_type,
        },
        "pagination": {"page_size": 10000},
    }
    response = requests.post(f"{url}/entries/archive/query", headers=headers, json=process_query)
    response.raise_for_status()

    entries = []
    for entry in response.json()["data"]:
        archive = entry["archive"]
        process_data = archive.get("data") or {}
        if "positon_in_experimental_plan" not in process_data:
            continue
        entry_type = (archive.get("metadata") or {}).get("entry_type") or process_data.get(
            "name", "Unknown Process"
        )
        entries.append((process_data, entry_type))

    entries.sort(key=lambda item: item[0]["positon_in_experimental_plan"])
    return entries


# entry types that are structural (the batch/sample records themselves), not
# a process or measurement someone would want in a cost catalog.
_EXCLUDED_SCHEMA_TYPES = {"HySprint_Batch", "HySprint_Sample"}


def discover_entry_types(url: str, token: str, sample_ids: list[str]) -> list[str]:
    """Every distinct schema (process AND measurement, e.g. HySprint_SpinCoating,
    HySprint_JVmeasurement) actually linked to these samples, via NOMAD's
    aggregation API - no hardcoded list of measurement types needed.
    Measurement schemas don't declare baseclasses.BaseMeasurement in
    section_defs the way process schemas declare baseclasses.BaseProcess
    (confirmed against live data: querying by that qualified name returns
    zero measurements), so this uses a plain entry_type aggregation instead,
    which works for anything.
    """
    headers = {"Authorization": f"Bearer {token}"}

    entry_id_query = {
        "required": {"metadata": "*"},
        "owner": "visible",
        "query": {"results.eln.lab_ids:any": sample_ids},
        "pagination": {"page_size": 10000},
    }
    response = requests.post(f"{url}/entries/query", headers=headers, json=entry_id_query)
    response.raise_for_status()
    entry_ids = [entry["entry_id"] for entry in response.json()["data"]]
    if not entry_ids:
        return []

    aggregation_query = {
        "owner": "visible",
        "query": {"entry_references.target_entry_id:any": entry_ids},
        "aggregations": {"entry_type_agg": {"terms": {"quantity": "entry_type", "size": 200}}},
        "pagination": {"page_size": 0},
    }
    response = requests.post(f"{url}/entries/query", headers=headers, json=aggregation_query)
    response.raise_for_status()
    terms = response.json().get("aggregations", {}).get("entry_type_agg", {}).get("terms", {})
    found_types = (entry["value"] for entry in terms.get("data", []))
    return sorted(
        entry_type for entry_type in found_types if entry_type not in _EXCLUDED_SCHEMA_TYPES
    )


# ---------------------------------------------------------------------------
# Cost math
# ---------------------------------------------------------------------------


def effective_cost_range(
    low: float | None, est: float | None, high: float | None
) -> tuple[float | None, float | None, float | None]:
    """Cost range fallback: if only a single (Est) value is filled in, that
    is the only value taken into account for both bounds."""
    effective_low = low if low is not None else est
    effective_high = high if high is not None else est
    return effective_low, est, effective_high


def cost_per_sample(cost_est: float | None, num_samples_covered: int) -> float | None:
    """A cost shared across a process instance's covered samples, divided
    evenly - the mechanism behind 'an evaporation run costs the same per
    batch whether it covers 1 or 16 samples'."""
    if cost_est is None or not num_samples_covered:
        return None
    return cost_est / num_samples_covered


# ---------------------------------------------------------------------------
# The shared cost reference file - single source of truth for cost figures
# ---------------------------------------------------------------------------
# An admin maintains one workbook (see DEFAULT_COST_REFERENCE_PATH below)
# with real prices/rates, marks them Verified, and saves it in place. Every
# user's export automatically reads that same file and carries forward
# anything that matches by name - no upload step, nothing for a regular
# user to configure.


@dataclass
class CostReference:
    material_prices: dict[str, dict] = field(default_factory=dict)  # by Material_Name
    labor_rates: dict[str, dict] = field(default_factory=dict)  # by Role (LABOR_ROLES)
    overhead_costs: dict[str, dict] = field(default_factory=dict)  # by Item

    @property
    def total_entries(self) -> int:
        return len(self.material_prices) + len(self.labor_rates) + len(self.overhead_costs)


def _header_index_map(ws) -> dict[str, int]:
    return {cell.value: index for index, cell in enumerate(ws[1]) if cell.value}


def parse_cost_reference_workbook(file_bytes: bytes) -> CostReference:
    """Read an LCC workbook's manually-entered cost columns, keyed for reuse
    in a fresh export. Works on either the lean admin cost-reference template
    (excel_export.build_cost_reference_template) or a full per-batch export -
    both use the same header names, looked up by name rather than fixed
    position. Only reads literal, non-formula columns - formula cells
    (Price_per_Gram_Est when computed from Price/Grams_on_Bottle, the
    Effective_*/Total_Price_Est/Summary cells) are recomputed fresh here in
    Python rather than trusting whatever Excel last cached, since openpyxl
    can't evaluate formulas itself.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    reference = CostReference()

    if "Materials" in workbook.sheetnames:
        ws = workbook["Materials"]
        idx = _header_index_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[idx["Material_Name"]]
            if not name:
                continue
            price = row[idx["Price"]] if "Price" in idx else None
            grams_on_bottle = row[idx["Grams_on_Bottle"]] if "Grams_on_Bottle" in idx else None
            if (
                isinstance(price, (int, float))
                and isinstance(grams_on_bottle, (int, float))
                and (grams_on_bottle)
            ):
                price_per_gram_est = price / grams_on_bottle
            else:
                literal = row[idx["Price_per_Gram_Est"]] if "Price_per_Gram_Est" in idx else None
                price_per_gram_est = literal if isinstance(literal, (int, float)) else None
            reference.material_prices[str(name).strip()] = {
                "cas_number": (row[idx["CAS_Number"]] if "CAS_Number" in idx else None) or None,
                "price_per_gram_est": price_per_gram_est,
                "verified": bool(row[idx["Verified"]]),
                "notes": row[idx["Notes"]] or "",
            }

    if "Labor" in workbook.sheetnames:
        ws = workbook["Labor"]
        idx = _header_index_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            role = row[idx["Role"]]
            if not role:
                continue
            reference.labor_rates[str(role).strip()] = {
                "hourly_rate_est": row[idx["Hourly_Rate_Est"]],
                "verified": bool(row[idx["Verified"]]),
            }

    if "Capital_Overhead_Disposal" in workbook.sheetnames:
        ws = workbook["Capital_Overhead_Disposal"]
        idx = _header_index_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = row[idx["Item"]]
            if not item:
                continue
            reference.overhead_costs[str(item).strip()] = {
                "cost_low": row[idx["Cost_Low"]],
                "cost_est": row[idx["Cost_Est"]],
                "cost_high": row[idx["Cost_High"]],
                "verified": bool(row[idx["Verified"]]),
                "notes": row[idx["Notes"]] or "",
            }

    return reference


# The single source of truth: one admin-maintained workbook living alongside
# the app's own code. Every user's session reads this same file automatically
# (no upload step) - an admin edits it in place (via git, or directly through
# NOMAD's upload file browser if this app is deployed as a North tool, since
# the app's working directory sits inside that same mounted upload) and the
# update is live for everyone on their next export. Mirrors the existing
# repo-root secrets.py convention (see hysprint_utils.access_token), just
# scoped to this app's own folder instead of the repo root.
DEFAULT_COST_REFERENCE_PATH = Path(__file__).resolve().parent / "cost_reference.xlsx"


def load_default_cost_reference(path: Path | None = None) -> CostReference | None:
    """Read the shared master cost file if it exists. Returns None (not an
    error) when the file is missing - a fresh install has no admin-entered
    costs yet, and the app should still work, just with blank costs."""
    reference_path = path if path is not None else DEFAULT_COST_REFERENCE_PATH
    if not reference_path.is_file():
        logger.info("No cost reference file at %s; costs will start blank.", reference_path)
        return None
    try:
        return parse_cost_reference_workbook(reference_path.read_bytes())
    except Exception:
        logger.exception("Could not read cost reference file at %s", reference_path)
        return None


# ---------------------------------------------------------------------------
# State container
# ---------------------------------------------------------------------------


class LCCDataManager:
    """Holds the extracted cost line items for the currently loaded batches.
    No widget imports, no global state - one instance per app session."""

    def __init__(self) -> None:
        self.batch_sample_counts: dict[str, int] = {}
        self.process_rows: list[ProcessStepRow] = []
        self.material_rows: list[MaterialRow] = []
        self.all_sample_ids: list[str] = []

    @property
    def has_data(self) -> bool:
        return bool(self.process_rows or self.material_rows)

    def load_batches(self, url: str, token: str, batch_ids: list[str]) -> None:
        self.batch_sample_counts = {}
        self.process_rows = []
        self.material_rows = []
        self.all_sample_ids = []

        samples_per_batch = fetch_samples_per_batch(url, token, batch_ids)
        sample_to_batch: dict[str, str] = {}
        for batch_id, sample_ids in samples_per_batch.items():
            self.batch_sample_counts[batch_id] = len(sample_ids)
            for sample_id in sample_ids:
                sample_to_batch[sample_id] = batch_id

        all_sample_ids = list(sample_to_batch)
        self.all_sample_ids = all_sample_ids
        if not all_sample_ids:
            logger.warning("No samples found in selected batches: %s", batch_ids)
            return

        process_entries = fetch_process_entries(url, token, all_sample_ids)
        for process_data, entry_type in process_entries:
            covered_sample_ids = [
                sample.get("lab_id")
                for sample in (process_data.get("samples") or [])
                if isinstance(sample, dict) and sample.get("lab_id")
            ]
            if not covered_sample_ids:
                continue

            # A process instance normally covers samples from a single batch;
            # if it spans more than one (e.g. a shared oven run across
            # batches), it is attributed here to the first covered sample's
            # batch as a documented simplification (see Guide sheet).
            batch_id = next(
                (sample_to_batch[sid] for sid in covered_sample_ids if sid in sample_to_batch),
                None,
            )
            if batch_id is None:
                continue

            num_samples_covered = len(covered_sample_ids)

            self.process_rows.extend(
                extract_process_step_rows(
                    process_data, entry_type, batch_id, num_samples_covered, covered_sample_ids
                )
            )
            self.material_rows.extend(
                extract_material_rows(
                    process_data, entry_type, batch_id, num_samples_covered, covered_sample_ids
                )
            )

        logger.info(
            "Loaded %d process rows, %d material rows for %d batches.",
            len(self.process_rows),
            len(self.material_rows),
            len(batch_ids),
        )

import base64
import io
import json
import logging
from datetime import datetime
from pathlib import Path

import ipywidgets as widgets
from IPython.display import display
from openpyxl import Workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)

try:
    from experiment_excel_builder import ExperimentExcelBuilder

    EXCEL_BUILDER_AVAILABLE = True
except ImportError:
    EXCEL_BUILDER_AVAILABLE = False
    logger.warning("ExperimentExcelBuilder not found. Install it or check the import path.")


class MinimalistExperimentBuilder:
    def __init__(self, templates_file="templates/process_templates.json"):
        self.templates_file = Path(templates_file)
        self.current_sequence = []
        self.templates = {}

        # Available process types
        processes = [
            "ALD",
            "Annealing",
            "Blade Coating",
            "Cleaning O2-Plasma",
            "Cleaning UV-Ozone",
            "Co-Evaporation",
            "Dip Coating",
            "Evaporation",
            "Generic Process",
            "Ink Recycling",
            "Inkjet Printing",
            "Laser Scribing",
            "Slot Die Coating",
            "Spin Coating",
            "Sputtering",
        ]
        self.available_processes = ["Experiment Info"] + sorted(processes)

        self.setup_widgets()
        self.load_templates()

        # Add default Experiment Info process
        self.current_sequence = [{"process": "Experiment Info"}]
        self._update_process_display()

    def load_templates(self):
        """Load templates from JSON file"""
        try:
            if self.templates_file.exists():
                with open(self.templates_file, "r") as f:
                    data = json.load(f)
                    self.templates = data.get("templates", {})

                logger.info("Loaded %d templates.", len(self.templates))
            else:
                logger.info("Creating default template file...")
                self.create_default_template()

            self._update_template_dropdown()

        except Exception as e:
            logger.error("Error loading templates: %s", e)
            self.create_default_template()

    def create_default_template(self):
        """Create a default template file"""
        default_data = {
            "metadata": {
                "version": "1.0",
                "description": "Process templates for laboratory automation",
                "last_updated": datetime.now().strftime("%Y-%m-%d"),
            },
            "templates": {
                "empty_template": {
                    "name": "Empty Template",
                    "description": "Start with just Experiment Info",
                    "category": "Basic",
                    "process_sequence": [{"process": "Experiment Info"}],
                },
                "test_process": {
                    "name": "Test Process",
                    "description": "Simple test process",
                    "category": "Test Processes",
                    "process_sequence": [
                        {"process": "Experiment Info"},
                        {
                            "process": "Spin Coating",
                            "config": {
                                "solvents": 2,
                                "solutes": 3,
                                "spinsteps": 1,
                                "antisolvent": True,
                            },
                        },
                        {"process": "Evaporation"},
                    ],
                },
                "simple_coating": {
                    "name": "Simple Coating",
                    "description": "Basic coating process",
                    "category": "Coating Processes",
                    "process_sequence": [
                        {"process": "Experiment Info"},
                        {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
                        {
                            "process": "Spin Coating",
                            "config": {"solvents": 1, "solutes": 1, "spinsteps": 1},
                        },
                        {"process": "Evaporation"},
                    ],
                },
            },
        }

        self.templates_file.parent.mkdir(parents=True, exist_ok=True)

        with open(self.templates_file, "w") as f:
            json.dump(default_data, f, indent=2)

        self.templates = default_data["templates"]
        logger.info("Created default template file with %d templates.", len(self.templates))
        self._update_template_dropdown()

    def setup_widgets(self):
        """Initialize all widgets"""
        # Template selection
        self.template_dropdown = widgets.Dropdown(
            options=["Select template..."],
            description="Template:",
            style={"description_width": "80px"},
            layout=widgets.Layout(width="300px"),
        )

        self.apply_button = widgets.Button(
            description="Apply", button_style="primary", layout=widgets.Layout(width="80px")
        )

        # Guide toggle button
        self.guide_toggle = widgets.Button(
            description="Show Guide",
            button_style="info",
            icon="question",
            layout=widgets.Layout(width="140px"),
        )

        # Guide content (loaded from file, initially hidden)
        self.guide_content = widgets.HTML(
            value=self._load_guide_content(), layout=widgets.Layout(display="none")
        )

        self.guide_visible = False

        # Process sequence area
        self.process_sequence_area = widgets.VBox([])

        # Generation controls
        self.is_testing_checkbox = widgets.Checkbox(
            value=True,
            description="Add example values on 1st row",
            style={"description_width": "initial"},
        )

        self.include_readme_checkbox = widgets.Checkbox(
            value=False,
            description="Create README.md for annotations",
            style={"description_width": "initial"},
        )

        self.generate_button = widgets.Button(
            description="Generate Excel",
            button_style="success",
            layout=widgets.Layout(width="150px"),
        )

        self.download_area = widgets.HTML(value="")
        self.status_output = widgets.Output()

        # Event handlers
        self.apply_button.on_click(self._on_apply_template)
        self.generate_button.on_click(self._on_generate_excel)
        self.guide_toggle.on_click(self._toggle_guide)

    def _load_guide_content(self):
        """Load guide content from external HTML file"""
        try:
            guide_file = Path("guide.html")
            if guide_file.exists():
                with open(guide_file, "r", encoding="utf-8") as f:
                    return f.read()
            else:
                # Fallback content if file doesn't exist
                return """
                <div style="padding: 15px; background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px; margin: 10px 0;">  # noqa: E501
                    <p>Guide file not found. Create 'guide.html' in your project directory.</p>
                </div>
                """
        except Exception as e:
            return f"""
            <div style="padding: 15px; background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px; margin: 10px 0;">  # noqa: E501
                <h4 style="color: #495057;">📖 Guide Error</h4>
                <p>Error loading guide: {e}</p>
            </div>
            """

    def _update_template_dropdown(self):
        """Update template dropdown with all available templates"""
        template_names = ["Select template..."] + [t["name"] for t in self.templates.values()]
        self.template_dropdown.options = template_names

    def _on_apply_template(self, button):
        """Apply selected template"""
        template_name = self.template_dropdown.value
        if template_name == "Select template...":
            self.status_output.clear_output()
            logger.warning("Please select a template.")
            return

        # Find template
        selected_template = None
        for template in self.templates.values():
            if template["name"] == template_name:
                selected_template = template
                break

        if selected_template:
            # Deep copy the process sequence
            self.current_sequence = []
            for process in selected_template["process_sequence"]:
                new_process = {"process": process["process"]}
                if "config" in process:
                    new_process["config"] = process["config"].copy()
                self.current_sequence.append(new_process)

            self._update_process_display()

            with self.status_output:
                self.status_output.clear_output()
                # print(f"✅ Applied: {template_name} ({len(self.current_sequence)} processes)")

    def _update_process_display(self):
        """Update the process sequence display"""
        if not self.current_sequence:
            self.process_sequence_area.children = []
            return

        process_rows = []

        for i, process_data in enumerate(self.current_sequence):
            row = self._create_process_row(i, process_data)
            process_rows.append(row)

        self.process_sequence_area.children = process_rows

    def _toggle_guide(self, button):
        """Toggle guide visibility"""
        self.guide_visible = not self.guide_visible

        if self.guide_visible:
            self.guide_content.layout.display = "block"
            self.guide_toggle.description = "Hide Guide"
            self.guide_toggle.icon = "times"
        else:
            self.guide_content.layout.display = "none"
            self.guide_toggle.description = "Show Guide"
            self.guide_toggle.icon = "question"

    def _create_process_row(self, index, process_data):
        """Create a single row for a process"""
        process_name = process_data["process"]
        config = process_data.get("config", {})

        # Index label
        index_label = widgets.HTML(
            value=f"<span style='font-weight: bold; color: #666; min-width: 25px; display: inline-block;'>{index + 1}.</span>",  # noqa: E501
            layout=widgets.Layout(width="30px"),
        )

        # Process type dropdown
        process_dropdown = widgets.Dropdown(
            options=self.available_processes,
            value=process_name,
            layout=widgets.Layout(width="180px"),
            disabled=(index == 0),  # Can't change Experiment Info
        )
        process_dropdown.observe(
            lambda change, idx=index: self._update_process_type(idx, change["new"]), names="value"
        )

        # Configuration controls
        numeric_controls, checkbox_controls = self._create_inline_config_controls(
            index, process_name, config
        )

        # Action buttons (always on the right)
        action_buttons = self._create_action_buttons(index)

        # Main row with numeric controls
        main_row_elements = [index_label, process_dropdown] + numeric_controls

        # Add spacer to push buttons to the right
        spacer = widgets.HTML(value="", layout=widgets.Layout(flex="1"))
        main_row_elements.append(spacer)
        main_row_elements.extend(action_buttons)

        main_row = widgets.HBox(
            main_row_elements,
            layout=widgets.Layout(
                margin="1px 0",
                padding="5px",
                border="1px solid #e0e0e0",
                border_radius="4px 4px 0 0" if checkbox_controls else "4px",
                align_items="center",
            ),
        )

        # If there are checkboxes, create a second row
        if checkbox_controls:
            checkbox_row_elements = [
                widgets.HTML(value="", layout=widgets.Layout(width="30px")),  # Align with index
                widgets.HTML(value="", layout=widgets.Layout(width="180px")),  # Align with dropdown
            ]
            checkbox_row_elements.extend(checkbox_controls)

            checkbox_row = widgets.HBox(
                checkbox_row_elements,
                layout=widgets.Layout(
                    margin="0",
                    padding="5px",
                    border="1px solid #e0e0e0",
                    border_top="none",
                    border_radius="0 0 4px 4px",
                    align_items="center",
                ),
            )

            # Return container with both rows
            return widgets.VBox([main_row, checkbox_row], layout=widgets.Layout(margin="2px 0"))
        else:
            return main_row

    def _create_inline_config_controls(self, index, process_name, config):
        """Create inline configuration controls - returns (numeric_controls, checkbox_controls)"""
        numeric_controls = []
        checkbox_controls = []

        # Add Atmospheric Values checkbox for all processes except Experiment Info
        # This must be added BEFORE the early return for non-configurable processes
        if process_name != "Experiment Info":
            atmospheric_checkbox = widgets.Checkbox(
                value=config.get("add_atmospheric", False),
                description="Add Atmospheric Values",
                style={"description_width": "initial"},
                layout=widgets.Layout(width="190px"),
            )
            atmospheric_checkbox.observe(
                lambda change, idx=index: self._update_config(
                    idx, "add_atmospheric", change["new"]
                ),
                names="value",
            )
            checkbox_controls.append(atmospheric_checkbox)

        # Check if process has configuration options
        configurable_processes = [
            "Spin Coating",
            "Cleaning O2-Plasma",
            "Cleaning UV-Ozone",
            "Inkjet Printing",
            "Co-Evaporation",
            "Ink Recycling",
            "Slot Die Coating",
            "Blade Coating",
        ]

        if process_name not in configurable_processes:
            return numeric_controls, checkbox_controls

        # Solvents
        if process_name in [
            "Spin Coating",
            "Cleaning O2-Plasma",
            "Cleaning UV-Ozone",
            "Inkjet Printing",
            "Ink Recycling",
            "Slot Die Coating",
            "Blade Coating",
        ]:
            solvents_widget = widgets.BoundedIntText(
                value=config.get("solvents", 0),
                min=0,
                max=20,
                description="Solvents:",
                style={"description_width": "55px"},
                layout=widgets.Layout(width="120px"),
            )
            solvents_widget.observe(
                lambda change, idx=index: self._update_config(idx, "solvents", change["new"]),
                names="value",
            )
            numeric_controls.append(solvents_widget)

        # Solutes
        if process_name in [
            "Spin Coating",
            "Inkjet Printing",
            "Ink Recycling",
            "Slot Die Coating",
            "Blade Coating",
        ]:
            solutes_widget = widgets.BoundedIntText(
                value=config.get("solutes", 0),
                min=0,
                max=20,
                description="Solutes:",
                style={"description_width": "50px"},
                layout=widgets.Layout(width="115px"),
            )
            solutes_widget.observe(
                lambda change, idx=index: self._update_config(idx, "solutes", change["new"]),
                names="value",
            )
            numeric_controls.append(solutes_widget)

        # Spin Steps
        if process_name == "Spin Coating":
            spinsteps_widget = widgets.BoundedIntText(
                value=config.get("spinsteps", 1),
                min=1,
                max=5,
                description="Steps:",
                style={"description_width": "40px"},
                layout=widgets.Layout(width="100px"),
            )
            spinsteps_widget.observe(
                lambda change, idx=index: self._update_config(idx, "spinsteps", change["new"]),
                names="value",
            )
            numeric_controls.append(spinsteps_widget)

        # Materials (for Co-Evaporation)
        if process_name == "Co-Evaporation":
            materials_widget = widgets.BoundedIntText(
                value=config.get("materials", 1),
                min=1,
                max=10,
                description="Materials:",
                style={"description_width": "65px"},
                layout=widgets.Layout(width="130px"),
            )
            materials_widget.observe(
                lambda change, idx=index: self._update_config(idx, "materials", change["new"]),
                names="value",
            )
            numeric_controls.append(materials_widget)

        # Precursors (for Ink Recycling)
        if process_name == "Ink Recycling":
            precursors_widget = widgets.BoundedIntText(
                value=config.get("precursors", 0),
                min=0,
                max=10,
                description="Precursors:",
                style={"description_width": "70px"},
                layout=widgets.Layout(width="135px"),
            )
            precursors_widget.observe(
                lambda change, idx=index: self._update_config(idx, "precursors", change["new"]),
                names="value",
            )
            numeric_controls.append(precursors_widget)

        # Checkboxes for Spin Coating
        if process_name == "Spin Coating":
            checkbox_options = [
                ("antisolvent", "Antisolvent"),
                ("gasquenching", "Gas Quenching"),
                ("vacuumquenching", "Vacuum Quenching"),
            ]

            for option_key, option_label in checkbox_options:
                checkbox = widgets.Checkbox(
                    value=config.get(option_key, False),
                    description=option_label,
                    style={"description_width": "initial"},
                    layout=widgets.Layout(width="140px"),
                )
                checkbox.observe(
                    lambda change, idx=index, key=option_key: self._update_config(
                        idx, key, change["new"]
                    ),
                    names="value",
                )
                checkbox_controls.append(checkbox)

        # Checkboxes for Spin Coating
        if process_name == "Blade Coating" or process_name == "Slot Die Coating":
            checkbox_options = [
                ("gasquenching", "Gas Quenching"),
                ("vacuumquenching", "Vacuum Quenching"),
            ]

            for option_key, option_label in checkbox_options:
                checkbox = widgets.Checkbox(
                    value=config.get(option_key, False),
                    description=option_label,
                    style={"description_width": "initial"},
                    layout=widgets.Layout(width="140px"),
                )
                checkbox.observe(
                    lambda change, idx=index, key=option_key: self._update_config(
                        idx, key, change["new"]
                    ),
                    names="value",
                )
                checkbox_controls.append(checkbox)

        # Checkboxes for Inkjet Printing
        if process_name == "Inkjet Printing":
            checkbox_options = [
                # ('annealing', 'Annealing'),
                ("gavd", "GAVD")
            ]

            for option_key, option_label in checkbox_options:
                checkbox = widgets.Checkbox(
                    value=config.get(option_key, False),
                    description=option_label,
                    style={"description_width": "initial"},
                    layout=widgets.Layout(width="100px"),
                )
                checkbox.observe(
                    lambda change, idx=index, key=option_key: self._update_config(
                        idx, key, change["new"]
                    ),
                    names="value",
                )
                checkbox_controls.append(checkbox)

        # Carbon paste for Evaporation
        if process_name == "Evaporation":
            checkbox = widgets.Checkbox(
                value=config.get("carbon_paste", False),
                description="Carbon Paste",
                style={"description_width": "initial"},
                layout=widgets.Layout(width="130px"),
            )
            checkbox.observe(
                lambda change, idx=index: self._update_config(idx, "carbon_paste", change["new"]),
                names="value",
            )
            checkbox_controls.append(checkbox)

        return numeric_controls, checkbox_controls

    def _create_action_buttons(self, index):
        """Create action buttons for a process row"""
        buttons = []

        # Add button (insert process below)
        add_button = widgets.Button(
            description="",
            icon="plus",
            button_style="success",
            layout=widgets.Layout(width="30px", height="28px"),
        )
        add_button.on_click(lambda b, idx=index: self._add_process_below(idx))
        buttons.append(add_button)

        # Remove button (can't remove Experiment Info)
        if index > 0:
            remove_button = widgets.Button(
                description="",
                icon="minus",
                button_style="danger",
                layout=widgets.Layout(width="30px", height="28px"),
            )
            remove_button.on_click(lambda b, idx=index: self._remove_process(idx))
            buttons.append(remove_button)
        else:
            # Spacer for Experiment Info row
            spacer = widgets.HTML(value="", layout=widgets.Layout(width="30px"))
            buttons.append(spacer)

        return buttons

    def _update_process_type(self, index, new_process_type):
        """Update process type"""
        if index < len(self.current_sequence):
            self.current_sequence[index]["process"] = new_process_type

            # Reset config when changing process type
            if new_process_type in [
                "Spin Coating",
                "Cleaning O2-Plasma",
                "Cleaning UV-Ozone",
                "Inkjet Printing",
                "Slot Die Coating",
                "Blade Coating",
            ]:
                self.current_sequence[index]["config"] = self._get_default_config(new_process_type)
            else:
                self.current_sequence[index].pop("config", None)

            self._update_process_display()

    def _get_default_config(self, process_name):
        """Get default configuration for a process"""
        defaults = {
            "Spin Coating": {
                "solvents": 1,
                "solutes": 1,
                "spinsteps": 1,
                "antisolvent": False,
                "gasquenching": False,
                "vacuumquenching": False,
            },
            "Blade Coating": {
                "solvents": 1,
                "solutes": 1,
                "gasquenching": False,
                "vacuumquenching": False,
            },
            "Cleaning O2-Plasma": {"solvents": 2},
            "Cleaning UV-Ozone": {"solvents": 2},
            "Inkjet Printing": {"solvents": 1, "solutes": 1, "annealing": False, "gavd": False},
            "Slot Die Coating": {
                "solvents": 1,
                "solutes": 1,
                "gasquenching": False,
                "vacuumquenching": False,
            },
            "Co-Evaporation": {"materials": 2},
            "Ink Recycling": {"solvents": 1, "solutes": 1, "precursors": 1},
            "Evaporation": {"carbon_paste": False},
        }
        return defaults.get(process_name, {})

    def _update_config(self, process_index, key, value):
        """Update configuration for a process"""
        if process_index < len(self.current_sequence):
            if "config" not in self.current_sequence[process_index]:
                self.current_sequence[process_index]["config"] = {}

            self.current_sequence[process_index]["config"][key] = value

    def _add_process_below(self, index):
        """Add a new process below the current one"""
        new_process = {"process": "Generic Process"}
        self.current_sequence.insert(index + 1, new_process)
        self._update_process_display()
        self.status_output.clear_output()
        logger.info("Added process at position %d.", index + 2)

    def _remove_process(self, index):
        """Remove a process (can't remove Experiment Info)"""
        if index > 0 and index < len(self.current_sequence):
            removed_process = self.current_sequence.pop(index)
            self._update_process_display()
            self.status_output.clear_output()
            logger.info("Removed: %s", removed_process["process"])

    def _on_generate_excel(self, button):
        """Generate Excel file"""
        if not self.current_sequence:
            self.status_output.clear_output()
            logger.warning("No processes configured.")
            return

        try:
            self.status_output.clear_output()
            logger.info("Generating Excel...")

            excel_data = self._generate_excel_data()

            logger.info("Final Excel data size: %d bytes", len(excel_data))

            if len(excel_data) > 0:
                if excel_data.startswith(b"PK"):
                    logger.info("Excel file signature is correct.")
                else:
                    logger.error("Not a valid Excel file! First bytes: %s", excel_data[:50])
            else:
                logger.error("Excel data is empty!")

            filename = f"experiment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            download_html = self._create_download_link(excel_data, filename)

            self.download_area.value = download_html
            logger.info("Download link created for: %s", filename)

            if self.include_readme_checkbox.value:
                readme_content = self._generate_readme()
                readme_filename = "README.md"
                readme_html = self._create_readme_download_link(readme_content, readme_filename)
                self.download_area.value += readme_html
                logger.info("README file created: %s", readme_filename)

        except Exception as e:
            self.status_output.clear_output()
            logger.exception("Error in _on_generate_excel: %s", e)

    def _generate_excel_data(self):
        """Generate Excel file using ExperimentExcelBuilder first, then openpyxl fallback"""

        if EXCEL_BUILDER_AVAILABLE:
            try:
                logger.info("Using ExperimentExcelBuilder (full detailed format)...")

                builder = ExperimentExcelBuilder(
                    self.current_sequence, self.is_testing_checkbox.value
                )
                builder.build_excel()

                buffer = io.BytesIO()
                builder.workbook.save(buffer)
                buffer.seek(0)
                excel_data = buffer.getvalue()

                logger.info("ExperimentExcelBuilder created file: %d bytes", len(excel_data))
                logger.info("Multiple sheets: %d worksheets", len(builder.workbook.worksheets))
                logger.info("Sheet names: %s", [ws.title for ws in builder.workbook.worksheets])

                if (
                    excel_data and len(excel_data) > 1000
                ):  # Should be much larger with detailed format
                    return excel_data
                else:
                    logger.warning(
                        "ExperimentExcelBuilder file seems too small, trying fallback..."
                    )

            except Exception as e:
                logger.error("ExperimentExcelBuilder failed: %s", e)
                logger.info("Falling back to basic openpyxl...")
        else:
            logger.warning("ExperimentExcelBuilder not available, using basic openpyxl...")

        # Fallback: Basic openpyxl (simple format)
        try:
            logger.info("Using basic openpyxl (simple format)...")

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Experiment Data"

            # Add content
            ws["A1"] = "Experiment Builder - Generated File"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A3"] = f"Testing Mode: {'Yes' if self.is_testing_checkbox.value else 'No'}"
            ws["A4"] = f"Number of Processes: {len(self.current_sequence)}"

            # Process sequence
            ws["A6"] = "Process Sequence:"
            ws["A6"].font = Font(bold=True)

            row = 7
            for i, process in enumerate(self.current_sequence, 1):
                ws[f"A{row}"] = f"{i}. {process['process']}"
                if "config" in process and process["config"]:
                    col = 2
                    for key, value in process["config"].items():
                        ws.cell(row=row, column=col, value=f"{key}: {value}")
                        col += 1
                row += 1

            logger.info("Basic Excel created: %d processes", len(self.current_sequence))

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            excel_data = buffer.getvalue()

            logger.info("Basic Excel file: %d bytes", len(excel_data))

            return excel_data

        except Exception as e:
            logger.exception("All Excel generation methods failed: %s", e)

            # Return error as text file
            error_content = f"""Excel Generation Error
=====================

Error: {e}
Timestamp: {datetime.now()}
Process Count: {len(self.current_sequence)}

Process Sequence:
"""
            for i, process in enumerate(self.current_sequence, 1):
                error_content += f"{i}. {process['process']}"
                if "config" in process and process["config"]:
                    config_str = ", ".join([f"{k}={v}" for k, v in process["config"].items()])
                    error_content += f" ({config_str})"
                error_content += "\n"

            return error_content.encode("utf-8")

    def _generate_readme(self):
        """Generate README.md template"""
        readme_template = """# Scientific Question
(What specific hypothesis or question does this experiment address?)

> Example: Does annealing temperature affect the VOC of triple-cation perovskite devices?

---

# Approach
(Briefly describe the experimental design, key parameters, and any deviations from standard protocol.)  # noqa: E501

## Conditions
| Parameter | Value |
|-----------|-------|
| ... | ... |

## Method
"""

        # Add process sequence
        for i, process in enumerate(self.current_sequence, 1):
            readme_template += f"{i}. {process['process']}"
            if "config" in process and process["config"]:
                config_str = ", ".join([f"{k}={v}" for k, v in process["config"].items()])
                readme_template += f" ({config_str})"
            readme_template += "\n"

        readme_template += """
---

# Results
(Summarize the key outputs. Link to data files or NOMAD entries where relevant.)

- **Key finding:** ...
- **Data location:** `path/to/data` or [NOMAD entry](#)

---

# Learnings
(What do the results tell you? What was surprising or confirmed?)

## Interpretation
...

## Caveats / Limitations
- ...

---

# Next Steps
(Concrete follow-up actions, not just observations.)
- [ ] Task one (owner, deadline)
- [ ] Task two

---

# References
- Related experiments: [link]
- Protocol used: [link]
"""

        return readme_template

    def _create_download_link(self, excel_data, filename):
        """Create download link"""
        b64_data = base64.b64encode(excel_data).decode()

        return f"""
        <div style="padding: 12px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 6px; margin: 10px 0;">  # noqa: E501
            <div style="display: flex; align-items: center; margin-bottom: 8px;">
                <span style="font-size: 18px; margin-right: 8px;">📊</span>
                <strong style="color: #155724;">Excel Ready</strong>
            </div>
            <div style="font-size: 14px; color: #155724; margin-bottom: 10px;">
                {filename} • {len(self.current_sequence)} processes • Testing: {"On" if self.is_testing_checkbox.value else "Off"}  # noqa: E501
            </div>
            <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_data}"   # noqa: E501
               download="{filename}" 
               style="display: inline-block; padding: 8px 16px; background-color: #28a745; color: white; text-decoration: none; border-radius: 4px; font-weight: bold; font-size: 14px;">  # noqa: E501
                📥 Download
            </a>
        </div>
        """

    def _create_readme_download_link(self, readme_content, filename):
        """Create download link for README file"""
        b64_data = base64.b64encode(readme_content.encode("utf-8")).decode()

        return f"""
        <div style="padding: 12px; background-color: #e7f3ff; border: 1px solid #b3d9ff; border-radius: 6px; margin: 10px 0;">  # noqa: E501
            <div style="display: flex; align-items: center; margin-bottom: 8px;">
                <span style="font-size: 18px; margin-right: 8px;">📝</span>
                <strong style="color: #004085;">README Template</strong>
            </div>
            <div style="font-size: 14px; color: #004085; margin-bottom: 10px;">
                {filename} • Experiment documentation template
            </div>
            <a href="data:text/markdown;base64,{b64_data}" 
               download="{filename}" 
               style="display: inline-block; padding: 8px 16px; background-color: #007bff; color: white; text-decoration: none; border-radius: 4px; font-weight: bold; font-size: 14px;">  # noqa: E501
                📥 Download README
            </a>
        </div>
        """

    def display(self):
        """Display the interface"""

        # Compact header
        header = widgets.HTML(
            value="""
            <div style="padding: 15px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 8px; margin-bottom: 20px;">  # noqa: E501
                <h2 style="margin: 0 0 5px 0; font-size: 1.8em;">🧪 Experiment Builder</h2>
                <p style="margin: 0; font-size: 0.95em; opacity: 0.9;">Configure processes and generate Excel files</p>  # noqa: E501
            </div>
            """
        )

        # Template selection - compact
        template_section = widgets.VBox(
            [
                widgets.HTML(
                    value="<h4 style='margin: 15px 0 8px 0; color: #2c3e50;'>📋 Template</h4>"
                ),
                widgets.HBox(
                    [self.template_dropdown, self.apply_button, self.guide_toggle],
                    layout=widgets.Layout(margin="0 0 15px 0"),
                ),
                self.guide_content,
            ]
        )

        # Process sequence - compact
        sequence_section = widgets.VBox(
            [
                widgets.HTML(
                    value="<h4 style='margin: 15px 0 8px 0; color: #2c3e50;'>⚙️ Process Sequence</h4>"  # noqa: E501
                ),
                self.process_sequence_area,
            ]
        )

        # Generation - compact
        generation_section = widgets.VBox(
            [
                widgets.HTML(
                    value="<h4 style='margin: 15px 0 8px 0; color: #2c3e50;'>📊 Generate</h4>"
                ),
                widgets.VBox(
                    [self.is_testing_checkbox, self.include_readme_checkbox],
                    layout=widgets.Layout(margin="0 0 10px 0"),
                ),
                widgets.HBox([self.generate_button], layout=widgets.Layout(margin="0 0 10px 0")),
                self.download_area,
            ]
        )

        # Status - compact
        status_section = widgets.VBox(
            [
                widgets.HTML(
                    value="<h5 style='margin: 15px 0 5px 0; color: #7f8c8d;'>📝 Status</h5>"
                ),
                self.status_output,
            ]
        )

        # Main interface - more compact
        main_interface = widgets.VBox(
            [header, template_section, sequence_section, generation_section, status_section],
            layout=widgets.Layout(padding="15px", max_width="1000px"),
        )

        # Only display once, don't return the interface to avoid duplication
        display(main_interface)


# Convenience function
def create_experiment_app(templates_file="templates/process_templates.json"):
    """Create and display the experiment builder interface"""
    app = MinimalistExperimentBuilder(templates_file)
    app.display()
    return app


def main():
    """Main function"""
    return create_experiment_app()

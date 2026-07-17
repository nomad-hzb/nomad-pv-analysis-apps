# data_manager.py
# Zero widget imports. Experiment state model, no-clobber write path, NOMAD live queries +
# session cache, and (in later implementation steps) Excel build/upload logic.

import io
import json
import logging
import statistics
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

import requests
from alias_config import resolve_progress_units
from experiment_excel_builder import ExperimentExcelBuilder
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
from sheet_experiment import add_experiment_sheet

from hysprint_utils.api_calls import (
    get_all_uploads,
    get_batch_ids,
    get_ids_in_batch,
    get_processing_steps,
)

logger = logging.getLogger(__name__)

MATERIAL_GATED_PROCESS_TYPES = {
    "Spin Coating",
    "Dip Coating",
    "Slot Die Coating",
    "Inkjet Printing",
    "Blade Coating",
    "Evaporation",
    "Sublimation",
    "Co-Evaporation",
    "Sputtering",
    "ALD",
}

MATERIAL_FIELD_KEY = "Material name"

# Experiment Info fields that vary per CHILD (pixel) row rather than per sample -
# see PixelFieldSpec. Confirmed against the real 20260603_Batch2028.xlsx.
PIXEL_FIELD_KEYS = {"Number of pixels", "Pixel area [cm^2]"}

# Experiment Info fields that are always derived/computed, never directly edited via
# ExperimentInfoPanel - "Variation" (matrix-computed), "Nomad ID"/"Sample" (derived at
# Excel-generation time from sample_number/child_index, see generate_full_workbook).
EXPERIMENT_INFO_COMPUTED_KEYS = {"Variation", "Nomad ID", "Sample"}

# Process catalog - mirrors Excel_creator's voila_experiment_app.py MinimalistExperimentBuilder
# (available_processes / configurable_processes / _get_default_config), reimplemented here as
# plain data since Excel_creator does not export these as importable module-level constants
# (they're literals inside a widget class) and this module must have zero widget imports.
# The actual field-generation logic (generate_steps_for_process) is NOT duplicated - see
# generate_header_workbook() below, which calls Excel_creator's own sheet_experiment.py.
AVAILABLE_PROCESSES = ["Experiment Info"] + sorted(
    [
        "ALD",
        "Annealing",
        "Blade Coating",
        "Cleaning O2-Plasma",
        "Cleaning UV-Ozone",
        "Co-Evaporation",
        "Dip Coating",
        "Evaporation",
        "Generic Process",
        "Ink Recycling",
        "Inkjet Printing",
        "Laser Scribing",
        "Slot Die Coating",
        "Spin Coating",
        "Sputtering",
    ]
)

CONFIGURABLE_PROCESS_TYPES = {
    "Spin Coating",
    "Cleaning O2-Plasma",
    "Cleaning UV-Ozone",
    "Inkjet Printing",
    "Co-Evaporation",
    "Ink Recycling",
    "Slot Die Coating",
    "Blade Coating",
}

DEFAULT_CONFIG_BY_PROCESS_TYPE: dict[str, dict] = {
    "Spin Coating": {
        "solvents": 1,
        "solutes": 1,
        "spinsteps": 1,
        "antisolvent": False,
        "gasquenching": False,
        "vacuumquenching": False,
    },
    "Blade Coating": {
        "solvents": 1,
        "solutes": 1,
        "gasquenching": False,
        "vacuumquenching": False,
    },
    "Cleaning O2-Plasma": {"solvents": 2},
    "Cleaning UV-Ozone": {"solvents": 2},
    "Inkjet Printing": {"solvents": 1, "solutes": 1, "annealing": False, "gavd": False},
    "Slot Die Coating": {
        "solvents": 1,
        "solutes": 1,
        "gasquenching": False,
        "vacuumquenching": False,
    },
    "Co-Evaporation": {"materials": 2},
    "Ink Recycling": {"solvents": 1, "solutes": 1, "precursors": 1},
    "Evaporation": {"carbon_paste": False},
}

# Declarative config-field catalog for the GUI layer: (config_key, label, applicable process
# types, min, max). Kept as data here so gui_components.py only has to render, not decide.
NUMERIC_CONFIG_FIELDS = [
    (
        "solvents",
        "Solvents",
        {
            "Spin Coating",
            "Cleaning O2-Plasma",
            "Cleaning UV-Ozone",
            "Inkjet Printing",
            "Ink Recycling",
            "Slot Die Coating",
            "Blade Coating",
        },
        0,
        20,
    ),
    (
        "solutes",
        "Solutes",
        {"Spin Coating", "Inkjet Printing", "Ink Recycling", "Slot Die Coating", "Blade Coating"},
        0,
        20,
    ),
    ("spinsteps", "Steps", {"Spin Coating"}, 1, 5),
    ("materials", "Materials", {"Co-Evaporation"}, 1, 10),
    ("precursors", "Precursors", {"Ink Recycling"}, 0, 10),
]

# (config_key, label, applicable process types). "add_atmospheric" applies to every real
# process (not Experiment Info, which never appears in ExperimentState.process_sequence).
BOOLEAN_CONFIG_FIELDS = [
    ("antisolvent", "Antisolvent", {"Spin Coating"}),
    ("gasquenching", "Gas Quenching", {"Spin Coating", "Blade Coating", "Slot Die Coating"}),
    ("vacuumquenching", "Vacuum Quenching", {"Spin Coating", "Blade Coating", "Slot Die Coating"}),
    ("gavd", "GAVD", {"Inkjet Printing"}),
    ("carbon_paste", "Carbon Paste", {"Evaporation"}),
]

ATMOSPHERIC_CONFIG_KEY = "add_atmospheric"


def default_config_for(process_type: str) -> dict:
    return dict(DEFAULT_CONFIG_BY_PROCESS_TYPE.get(process_type, {}))


def _is_filled(value: Any) -> bool:
    return value is not None and value != ""


class FieldProvenance(BaseModel):
    source: Literal["manual", "batch_template", "process_override", "computed"] = "manual"
    source_batch_id: str | None = None
    source_sample_id: str | None = None


class ProcessFieldSpec(BaseModel):
    key: str
    varies: bool = False
    alias_group: str | None = None
    value: Any = None
    provenance: FieldProvenance | None = None
    is_outlier: bool = False
    # Per-field opt-out of the completion count/nudge review - distinct from
    # ProcessInstance.counts_toward_progress() (process-level material-gating). Defaults
    # to True (today's behavior: every field counts) - the product ask was to let users
    # exclude fields that "aren't really important" (e.g. Notes, Tool/GB name) from the
    # completion bar without having to fill them just to make the number look right.
    required_for_progress: bool = True
    per_sample_values: dict[int, Any] = Field(default_factory=dict)
    per_sample_provenance: dict[int, FieldProvenance] = Field(default_factory=dict)

    def is_filled(self) -> bool:
        if self.varies:
            return any(_is_filled(v) for v in self.per_sample_values.values())
        return _is_filled(self.value)


class PixelFieldSpec(BaseModel):
    """Number of pixels / Pixel area - vary per CHILD row, not per sample."""

    key: str
    varies: bool = False
    value: Any = None
    per_child_values: dict[tuple[int, int], Any] = Field(default_factory=dict)

    def is_filled(self) -> bool:
        if self.varies:
            return any(_is_filled(v) for v in self.per_child_values.values())
        return _is_filled(self.value)


class ProcessInstance(BaseModel):
    process_type: str
    sequence_index: int
    config: dict = Field(default_factory=dict)
    field_specs: dict[str, ProcessFieldSpec] = Field(default_factory=dict)
    source_override_batch_id: str | None = None

    def is_material_gated(self) -> bool:
        return self.process_type in MATERIAL_GATED_PROCESS_TYPES

    def counts_toward_progress(self) -> bool:
        if not self.is_material_gated():
            return True
        material_field = self.field_specs.get(MATERIAL_FIELD_KEY)
        return material_field is not None and material_field.is_filled()


class SamplePlan(BaseModel):
    variation_group_index: int
    sample_number: int
    child_count: int = 0


class ExperimentState(BaseModel):
    process_sequence: list[ProcessInstance] = Field(default_factory=list)
    experiment_info_fields: dict[str, ProcessFieldSpec] = Field(default_factory=dict)
    pixel_fields: dict[str, PixelFieldSpec] = Field(default_factory=dict)
    samples: list[SamplePlan] = Field(default_factory=list)
    whole_experiment_template_batch_id: str | None = None

    def get_process(self, sequence_index: int) -> ProcessInstance:
        for process in self.process_sequence:
            if process.sequence_index == sequence_index:
                return process
        raise KeyError(f"No process at sequence_index {sequence_index}")

    def renumber_sequence_indices(self) -> None:
        """Experiment Info always occupies index 0 (implicitly); real processes start at 1,
        matching sheet_experiment.py's incremental_number numbering."""
        for position, process in enumerate(self.process_sequence, start=1):
            process.sequence_index = position

    def add_process(
        self, process_type: str, config: dict | None = None, at_index: int | None = None
    ) -> ProcessInstance:
        process = ProcessInstance(process_type=process_type, sequence_index=0, config=config or {})
        if at_index is None:
            self.process_sequence.append(process)
        else:
            self.process_sequence.insert(at_index, process)
        self.renumber_sequence_indices()
        return process

    def remove_process(self, sequence_index: int) -> None:
        self.process_sequence = [
            p for p in self.process_sequence if p.sequence_index != sequence_index
        ]
        self.renumber_sequence_indices()

    def sample_numbers(self) -> list[int]:
        return [s.sample_number for s in self.samples]

    def add_sample(
        self, variation_group_index: int, sample_number: int | None = None, child_count: int = 0
    ) -> SamplePlan:
        """variation_group_index is setup-time grouping only (uneven counts per group are
        expected, e.g. group A = 3 samples, group B = 5) - distinct from the auto-computed
        Variation label, which is always derived fresh from checked varying fields."""
        if sample_number is None:
            sample_number = max((s.sample_number for s in self.samples), default=0) + 1
        plan = SamplePlan(
            variation_group_index=variation_group_index,
            sample_number=sample_number,
            child_count=child_count,
        )
        self.samples.append(plan)
        return plan

    def remove_sample(self, sample_number: int) -> None:
        self.samples = [s for s in self.samples if s.sample_number != sample_number]


def compute_sample_set_split(total_samples: int, num_sets: int) -> list[int]:
    """Divides total_samples as evenly as possible across num_sets, front-loading the
    remainder (e.g. 15 samples / 4 sets -> [4, 4, 4, 3]) - the 'most natural division'
    preloaded into SampleSetupPanel's per-set count inputs before the user adjusts them."""
    if num_sets <= 0:
        return []
    base, remainder = divmod(max(total_samples, 0), num_sets)
    return [base + 1 if i < remainder else base for i in range(num_sets)]


def set_field_if_empty(
    spec: ProcessFieldSpec,
    value: Any,
    provenance: FieldProvenance | None = None,
    sample_number: int | None = None,
) -> bool:
    """Write autofilled/derived data. Never overwrites an existing value. Returns True
    if a write happened."""
    if spec.varies:
        if sample_number is None:
            raise ValueError("sample_number is required when spec.varies is True")
        if _is_filled(spec.per_sample_values.get(sample_number)):
            return False
        spec.per_sample_values[sample_number] = value
        if provenance is not None:
            spec.per_sample_provenance[sample_number] = provenance
        return True

    if _is_filled(spec.value):
        return False
    spec.value = value
    if provenance is not None:
        spec.provenance = provenance
    return True


def set_field_manual(
    spec: ProcessFieldSpec,
    value: Any,
    sample_number: int | None = None,
) -> None:
    """Write a direct user edit. Always overwrites, clears any outlier flag."""
    if spec.varies:
        if sample_number is None:
            raise ValueError("sample_number is required when spec.varies is True")
        spec.per_sample_values[sample_number] = value
        spec.per_sample_provenance[sample_number] = FieldProvenance(source="manual")
    else:
        spec.value = value
        spec.provenance = FieldProvenance(source="manual")
    spec.is_outlier = False


def set_field_varies(spec: ProcessFieldSpec, varies: bool, sample_numbers: list[int]) -> None:
    """Toggle a field's scope. Turning varies on seeds any currently-empty per-sample
    slots from the existing constant value (no-clobber). Turning it off never destroys
    per_sample_values, so re-enabling later restores prior entries."""
    if varies and not spec.varies and _is_filled(spec.value):
        for sample_number in sample_numbers:
            if not _is_filled(spec.per_sample_values.get(sample_number)):
                spec.per_sample_values[sample_number] = spec.value
                if spec.provenance is not None:
                    spec.per_sample_provenance[sample_number] = spec.provenance
    spec.varies = varies


def set_field_required_for_progress(spec: ProcessFieldSpec, required: bool) -> None:
    """Toggle whether this field counts toward the completion bar / nudge review. Pure
    display-preference flag - never touches the field's value or provenance, and does
    not affect Excel generation (a "not required" field is still written to the output
    file if it has a value, same as any other field)."""
    spec.required_for_progress = required


def set_pixel_field_if_empty(
    spec: PixelFieldSpec,
    value: Any,
    sample_number: int,
    child_index: int,
) -> bool:
    if spec.varies:
        key = (sample_number, child_index)
        if _is_filled(spec.per_child_values.get(key)):
            return False
        spec.per_child_values[key] = value
        return True
    if _is_filled(spec.value):
        return False
    spec.value = value
    return True


def set_pixel_field_manual(
    spec: PixelFieldSpec,
    value: Any,
    sample_number: int,
    child_index: int,
) -> None:
    if spec.varies:
        spec.per_child_values[(sample_number, child_index)] = value
    else:
        spec.value = value


# ---------------------------------------------------------------------------
# Excel header generation - reuses Excel_creator's sheet_experiment.py exactly, then
# reconstructs a column map by reading back the generated header instead of duplicating
# generate_steps_for_process (which is a nested function, not exported).
# ---------------------------------------------------------------------------


def process_sequence_to_dicts(state: ExperimentState) -> list[dict]:
    """Build the [{"process": ..., "config": {...}}, ...] shape sheet_experiment.py expects."""
    sequence = [{"process": "Experiment Info"}]
    for process in state.process_sequence:
        entry: dict = {"process": process.process_type}
        if process.config:
            entry["config"] = dict(process.config)
        sequence.append(entry)
    return sequence


def generate_header_workbook(state: ExperimentState) -> Workbook:
    """Row 1 (process labels) + row 2 (field labels) skeleton only - no sample data rows.
    Always called with is_testing=False; smart_databaser writes real data rows itself,
    never Excel_creator's fabricated is_testing row."""
    workbook = Workbook()
    add_experiment_sheet(workbook, process_sequence_to_dicts(state), is_testing=False)
    return workbook


def _parse_sequence_index(row1_label: str) -> int:
    if row1_label == "Experiment Info":
        return 0
    prefix = row1_label.split(":", 1)[0]
    return int(prefix)


def build_column_map(worksheet: Worksheet) -> dict[tuple[int, str], int]:
    """{(sequence_index, field_key): column_index}, read back from row 1 + row 2 rather than
    reimplementing sheet_experiment.py's column layout. Robust to header quirks such as a
    process block whose row 1 label wasn't merged across its columns (observed in the real
    20260603_Batch2028.xlsx) since it only depends on cell values, not merged ranges."""
    column_map: dict[tuple[int, str], int] = {}
    current_sequence_index: int | None = None
    for col in range(1, worksheet.max_column + 1):
        row1_value = worksheet.cell(row=1, column=col).value
        if row1_value is not None:
            current_sequence_index = _parse_sequence_index(row1_value)
        if current_sequence_index is None:
            continue
        field_key = worksheet.cell(row=2, column=col).value
        if field_key is None:
            continue
        column_map[(current_sequence_index, field_key)] = col
    return column_map


def sync_field_specs_from_columns(
    state: ExperimentState, column_map: dict[tuple[int, str], int]
) -> None:
    """Additive only: creates a ProcessFieldSpec/PixelFieldSpec for every column that
    doesn't have one yet. Never removes an existing spec (e.g. after reducing a solvent
    count), so already-filled values are never silently dropped."""
    for sequence_index, field_key in column_map:
        if sequence_index == 0:
            if field_key in PIXEL_FIELD_KEYS:
                if field_key not in state.pixel_fields:
                    state.pixel_fields[field_key] = PixelFieldSpec(key=field_key)
                continue
            bucket = state.experiment_info_fields
        else:
            bucket = state.get_process(sequence_index).field_specs
        if field_key not in bucket:
            bucket[field_key] = ProcessFieldSpec(key=field_key)


def rebuild_field_specs(state: ExperimentState) -> dict[tuple[int, str], int]:
    """Regenerate the header workbook from the current process sequence/config, additively
    sync field_specs/pixel_fields, and return the fresh column map. Call after any
    add/remove/config-change to the process sequence."""
    workbook = generate_header_workbook(state)
    worksheet = workbook.active
    column_map = build_column_map(worksheet)
    sync_field_specs_from_columns(state, column_map)
    return column_map


# ---------------------------------------------------------------------------
# NOMAD live value sourcing - session-scoped cache, per-process-type field mapping.
#
# The field paths themselves live in config/field_mappings.json, not here, specifically
# so unit_verified flags can be flipped and new process types/fields added without
# touching this module - see that file's "_readme" key for the schema and the unit-
# verification caveat (the archive query returns no unit metadata, so numeric fields are
# copied unconverted until confirmed against the NOMAD web GUI).
# ---------------------------------------------------------------------------

FIELD_MAPPINGS_CONFIG_PATH = Path(__file__).parent / "config" / "field_mappings.json"


def _get_path(data: Any, path: list) -> Any:
    current = data
    for key in path:
        if current is None:
            return None
        try:
            current = current[key] if isinstance(key, int) else current.get(key)
        except (IndexError, KeyError, TypeError):
            return None
    return current


def _get_path_any(data: Any, paths: list[list]) -> Any:
    """Tries each alternative path in order, returns the first non-None hit. Every
    loaded field entry is normalized to a list of paths (see load_field_mappings), even
    when only one applies, so this is the only path-resolution function callers need."""
    for path in paths:
        value = _get_path(data, path)
        if value is not None:
            return value
    return None


def _resolve_indexed_path_template(path_template: list, index: int) -> list:
    return [index if segment == "{i}" else segment for segment in path_template]


def load_field_mappings(
    config_path: Path | None = None,
) -> dict[str, dict[str, tuple[list[list], bool]]]:
    """Loads config/field_mappings.json into {process_type: {excel_field_key: (paths,
    unit_verified)}}. Pass config_path to load an alternate file (e.g. in tests).

    A field entry's 'path' (the common case) is a single archive path; an entry may
    instead use 'paths' (plural) - a list of alternative full paths, tried in order by
    _get_path_any, first non-None wins. This is for cases where the archive stores
    equivalent data under a different parent key depending on other data on the same
    step - e.g. Evaporation's field values live under organic_evaporation,
    inorganic_evaporation, or perovskite_evaporation depending on an 'Organic'
    checkbox/co-evaporation flag not itself tracked by this app, but the field names
    inside each are identical. Every entry is normalized here to a list of paths (even
    single-path ones), so callers never special-case either shape. Same 'path_template'
    / 'path_templates' (plural) split for indexed_fields."""
    path = config_path or FIELD_MAPPINGS_CONFIG_PATH
    with open(path, encoding="utf-8") as config_file:
        raw = json.load(config_file)

    mappings: dict[str, dict[str, tuple[list[list], bool]]] = {}
    for process_type, spec in raw.get("process_types", {}).items():
        field_paths: dict[str, tuple[list[list], bool]] = {}
        for excel_key, field_spec in spec.get("fields", {}).items():
            paths = field_spec["paths"] if "paths" in field_spec else [field_spec["path"]]
            field_paths[excel_key] = (paths, field_spec["unit_verified"])
        for indexed in spec.get("indexed_fields", []):
            start, end = indexed["range"]
            templates = (
                indexed["path_templates"]
                if "path_templates" in indexed
                else [indexed["path_template"]]
            )
            for n in range(start, end + 1):
                excel_key = indexed["excel_key_template"].format(n=n)
                resolved_paths = [
                    _resolve_indexed_path_template(template, n - 1) for template in templates
                ]
                field_paths[excel_key] = (resolved_paths, indexed["unit_verified"])
        mappings[process_type] = field_paths
    return mappings


# Process-type -> {excel_field_key: (json_path, unit_verified)}, loaded from
# config/field_mappings.json at import time. Extend by editing that file, not this one.
PROCESS_TYPE_FIELD_PATHS: dict[str, dict[str, tuple[list, bool]]] = load_field_mappings()


def load_indexed_config_keys(config_path: Path | None = None) -> dict[str, dict[str, str]]:
    """{process_type: {config_key: excel_key_template}} - which indexed_fields entries in
    field_mappings.json correspond to which process config count (e.g. Spin Coating's
    "solvents" config maps to the "Solvent {n} name" indexed field's config_key tag) -
    used by infer_config_from_source_step to widen a target process's config to match
    what a source batch step actually has data for."""
    path = config_path or FIELD_MAPPINGS_CONFIG_PATH
    with open(path, encoding="utf-8") as config_file:
        raw = json.load(config_file)

    result: dict[str, dict[str, str]] = {}
    for process_type, spec in raw.get("process_types", {}).items():
        mapping = {
            indexed["config_key"]: indexed["excel_key_template"]
            for indexed in spec.get("indexed_fields", [])
            if indexed.get("config_key")
        }
        if mapping:
            result[process_type] = mapping
    return result


# Process-type -> {config_key: excel_key_template}, loaded from config/field_mappings.json
# alongside PROCESS_TYPE_FIELD_PATHS above.
INDEXED_CONFIG_KEYS: dict[str, dict[str, str]] = load_indexed_config_keys()


def infer_config_from_source_step(process_type: str, step: dict) -> dict:
    """Best-effort config (solvents/solutes/spinsteps/... counts) inferred from how many
    of a source step's indexed fields (Solvent N name, Rotation speed N, ...) actually
    have data, per INDEXED_CONFIG_KEYS - so a target process autofilled from this step
    gets a field_spec slot for every value the source can supply, instead of silently
    dropping values beyond whatever count the target happened to already be configured
    for (e.g. a source with 2 solvents but a target still configured for 1)."""
    field_paths = PROCESS_TYPE_FIELD_PATHS.get(process_type, {})
    inferred: dict[str, int] = {}
    for config_key, excel_key_template in INDEXED_CONFIG_KEYS.get(process_type, {}).items():
        count = 0
        for n in range(1, 6):
            path_info = field_paths.get(excel_key_template.format(n=n))
            if path_info is not None and _get_path_any(step, path_info[0]) is not None:
                count = n
        if count:
            inferred[config_key] = count
    return inferred


class NomadSessionCache:
    """Session-scoped, in-memory-only cache - never persisted to disk, never shared across
    sessions. Mirrors 0Global_analyzer's DataManager plain-dict-cache pattern; deliberately
    does NOT use hysprint_utils.api_calls.init_cache() (a requests_cache disk cache)."""

    def __init__(self) -> None:
        self._sample_ids_by_batch: dict[str, list[str]] = {}
        self._processing_steps_by_batch: dict[str, list[dict]] = {}
        self._batch_ids: list[str] | None = None
        self._uploads: list[dict] | None = None

    def clear(self) -> None:
        self._sample_ids_by_batch.clear()
        self._processing_steps_by_batch.clear()
        self._batch_ids = None
        self._uploads = None

    def get_batch_ids(self, url: str, token: str) -> list[str]:
        """Shared by every batch picker widget (whole-experiment template, each
        per-process override) so opening several pickers in one session costs one
        get_batch_ids() call, not one per picker."""
        if self._batch_ids is None:
            self._batch_ids = get_batch_ids(url, token)
        return self._batch_ids

    def get_uploads(self, url: str, token: str) -> list[dict]:
        """The user's own already-created NOMAD uploads, for the upload-target picker."""
        if self._uploads is None:
            self._uploads = get_all_uploads(url, token)
        return self._uploads

    def get_sample_ids(self, url: str, token: str, batch_id: str) -> list[str]:
        if batch_id not in self._sample_ids_by_batch:
            self._sample_ids_by_batch[batch_id] = get_ids_in_batch(url, token, [batch_id])
        return self._sample_ids_by_batch[batch_id]

    def get_processing_steps(self, url: str, token: str, batch_id: str) -> list[dict]:
        if batch_id not in self._processing_steps_by_batch:
            sample_ids = self.get_sample_ids(url, token, batch_id)
            self._processing_steps_by_batch[batch_id] = get_processing_steps(url, token, sample_ids)
        return self._processing_steps_by_batch[batch_id]

    def peek_processing_steps(self, batch_id: str) -> list[dict] | None:
        """Cache-only lookup (no network call) - returns None if this batch's steps have
        not been fetched yet this session. Used for live preview placeholders, which must
        never trigger a fetch merely by rendering a field row."""
        return self._processing_steps_by_batch.get(batch_id)


# Real NOMAD archive 'method' strings that don't match this app's AVAILABLE_PROCESSES
# labels 1:1, verified against real batch HZB_JJ_1_A on 2026-07-15. Extend this (not the
# resolver function below) when a new mismatch is found.
_METHOD_ALIASES: dict[str, str] = {
    "Atomic Layer Deposition": "ALD",
}


def _has_any_data(entries: Any) -> bool:
    return isinstance(entries, list) and any(
        isinstance(entry, dict) and any(_is_filled(v) for v in entry.values()) for entry in entries
    )


def resolve_process_type(step: dict) -> str | None:
    """Translates a source step's raw 'method' into this app's own AVAILABLE_PROCESSES
    label, so autofill/occurrence-listing/sequence-replication all agree on which app
    process type a given source step maps to. Returns None if unrecognized.

    Two real mismatches verified against batch HZB_JJ_1_A: (1) simple renames, handled
    via _METHOD_ALIASES (e.g. real "Atomic Layer Deposition" vs this app's "ALD"); (2) a
    single real "Cleaning" step covers what this app models as two separate process
    types ("Cleaning O2-Plasma" / "Cleaning UV-Ozone") - disambiguated by which of
    cleaning_uv/cleaning_plasma actually has data on that step (both empty defaults to
    "Cleaning UV-Ozone", an arbitrary but documented choice; if both have data,
    UV-Ozone also wins, since Excel_creator can't represent 'both' as one process
    instance either)."""
    method = step.get("method")
    if method is None:
        return None
    if method in AVAILABLE_PROCESSES:
        return method
    if method in _METHOD_ALIASES:
        return _METHOD_ALIASES[method]
    if method == "Cleaning":
        uv_has_data = _has_any_data(step.get("cleaning_uv"))
        plasma_has_data = _has_any_data(step.get("cleaning_plasma"))
        if plasma_has_data and not uv_has_data:
            return "Cleaning O2-Plasma"
        return "Cleaning UV-Ozone"
    return None


def steps_for_process_type(steps: list[dict], process_type: str) -> list[dict]:
    """get_processing_steps() is already sorted by positon_in_experimental_plan; matches
    via resolve_process_type() rather than a raw 'method' comparison, since the archive's
    method string doesn't always match this app's process type labels 1:1."""
    return [s for s in steps if resolve_process_type(s) == process_type]


def fetch_process_field_values(
    url: str,
    token: str,
    cache: NomadSessionCache,
    batch_id: str,
    process_type: str,
    occurrence: int = 0,
) -> tuple[dict[str, Any], str | None]:
    """Values for the `occurrence`-th step of process_type in the batch (0 = earliest
    position, matching sort order), plus the lab_id of its first referenced sample for
    provenance tagging. Returns ({}, None) if the batch has no matching step."""
    field_paths = PROCESS_TYPE_FIELD_PATHS.get(process_type)
    if not field_paths:
        return {}, None
    steps = steps_for_process_type(cache.get_processing_steps(url, token, batch_id), process_type)
    if occurrence >= len(steps):
        return {}, None
    step = steps[occurrence]
    values = {}
    for field_key, (paths, _unit_verified) in field_paths.items():
        value = _get_path_any(step, paths)
        if value is not None:
            values[field_key] = value
    samples = step.get("samples") or []
    source_sample_id = samples[0]["lab_id"] if samples else None
    return values, source_sample_id


def list_process_occurrences(
    url: str, token: str, cache: NomadSessionCache, batch_id: str, process_type: str
) -> list[tuple[int, str]]:
    """(occurrence_index, label) for every step of process_type in batch_id, in the same
    order autofill_process_from_batch's `occurrence` parameter indexes into. Label is the
    step's Material name when process_type has that field mapped (the material-gated
    process types), else a generic 'Occurrence N' fallback - lets the 'adopt from
    template batch' picker distinguish multiple same-type steps (e.g. several Spin
    Coating layers) by what they actually deposited."""
    steps = steps_for_process_type(cache.get_processing_steps(url, token, batch_id), process_type)
    material_path_info = PROCESS_TYPE_FIELD_PATHS.get(process_type, {}).get(MATERIAL_FIELD_KEY)
    occurrences = []
    for index, step in enumerate(steps):
        label = None
        if material_path_info is not None:
            material_value = _get_path_any(step, material_path_info[0])
            if material_value:
                label = str(material_value)
        occurrences.append((index, label or f"Occurrence {index + 1}"))
    return occurrences


def autofill_process_from_batch(
    process: ProcessInstance,
    url: str,
    token: str,
    cache: NomadSessionCache,
    batch_id: str,
    occurrence: int = 0,
    provenance_source: Literal["batch_template", "process_override"] = "batch_template",
) -> int:
    """Writes every mapped field found in the batch's matching process step via
    set_field_if_empty (no-clobber), and flags each written value as an outlier if it
    differs substantially from the field's historical distribution at the same
    experimental-plan position (see compute_field_distribution_for_occurrence) - this is
    what feeds the outlier-flagged half of the nudge queue and the provenance tag's red
    coloring in the GUI. Returns how many fields were actually written."""
    values, source_sample_id = fetch_process_field_values(
        url, token, cache, batch_id, process.process_type, occurrence
    )
    written = 0
    for field_key, value in values.items():
        spec = process.field_specs.get(field_key)
        if spec is None:
            continue
        provenance = FieldProvenance(
            source=provenance_source, source_batch_id=batch_id, source_sample_id=source_sample_id
        )
        if set_field_if_empty(spec, value, provenance):
            written += 1
            distribution = compute_field_distribution_for_occurrence(
                url,
                token,
                cache,
                batch_id,
                process.process_type,
                field_key,
                occurrence,
                exclude_occurrence=True,
            )
            spec.is_outlier = is_outlier(value, distribution)
    return written


def occurrence_index_for_process(state: ExperimentState, process: ProcessInstance) -> int:
    """How many processes of the same process_type precede this one in the sequence -
    used to source the Nth occurrence of that type from a template/override batch (e.g.
    the target sequence's 2nd Slot Die Coating sources from the batch's 2nd Slot Die
    Coating step, not always the first)."""
    count = 0
    for candidate in state.process_sequence:
        if candidate.sequence_index == process.sequence_index:
            return count
        if candidate.process_type == process.process_type:
            count += 1
    return count


def preview_value_for_field(
    state: ExperimentState, process: ProcessInstance, field_key: str, cache: NomadSessionCache
) -> Any:
    """Best-effort preview of what field_key WOULD become if re-autofilled from this
    process's active source batch (its override, else the whole-experiment template) -
    read from already-cached NOMAD data only, so it never triggers a network call just
    from rendering a field row. Returns None if there's no active source, its steps
    aren't cached yet, or the source has no value for this field. Purely a display hint
    (see gui_components' Text.placeholder usage) - never written via set_field_if_empty."""
    batch_id = process.source_override_batch_id or state.whole_experiment_template_batch_id
    if not batch_id:
        return None
    steps = cache.peek_processing_steps(batch_id)
    if steps is None:
        return None
    path_info = PROCESS_TYPE_FIELD_PATHS.get(process.process_type, {}).get(field_key)
    if path_info is None:
        return None
    matching = steps_for_process_type(steps, process.process_type)
    occurrence = occurrence_index_for_process(state, process)
    if occurrence >= len(matching):
        return None
    return _get_path_any(matching[occurrence], path_info[0])


def expand_process_config_for_source(
    state: ExperimentState,
    process: ProcessInstance,
    url: str,
    token: str,
    cache: NomadSessionCache,
    batch_id: str,
    occurrence: int,
) -> None:
    """Widens process.config (solvents/solutes/spinsteps/... counts) to at least cover
    what the source batch's matching step actually has data for, then regenerates
    field_specs (additive only, see rebuild_field_specs) so autofill_process_from_batch
    has a slot to write every value the source can supply. Never shrinks an
    already-larger config. No-op if the occurrence doesn't exist in the source batch."""
    all_steps = cache.get_processing_steps(url, token, batch_id)
    steps = steps_for_process_type(all_steps, process.process_type)
    if occurrence >= len(steps):
        return
    inferred = infer_config_from_source_step(process.process_type, steps[occurrence])
    changed = False
    for key, count in inferred.items():
        if count > process.config.get(key, 0):
            process.config[key] = count
            changed = True
    if changed:
        rebuild_field_specs(state)


def _distinct_positions_in_order(steps: list[dict]) -> list[dict]:
    """One representative step per distinct positon_in_experimental_plan, in the order
    they first appear (get_processing_steps is already sorted by position) - collapses
    the per-sample/per-variation-group duplicate entries get_processing_steps returns for
    the same logical sequence step (e.g. several samples annealed at different
    temperatures but at the same sequence position) down to one, taking the first-listed
    entry (matches 'if there are variations, always grab the first value')."""
    seen_positions: set = set()
    representatives = []
    for step in steps:
        position = step.get("positon_in_experimental_plan")
        if position in seen_positions:
            continue
        seen_positions.add(position)
        representatives.append(step)
    return representatives


def build_process_sequence_from_batch(
    url: str, token: str, cache: NomadSessionCache, batch_id: str
) -> list[ProcessInstance]:
    """One ProcessInstance per distinct sequence position in batch_id (see
    _distinct_positions_in_order), in position order, each with its config sized to match
    what that step actually has data for (infer_config_from_source_step). Powers "Copy
    Values From Batch": the whole-experiment template replicates the source batch's own
    process sequence shape, not just fills values into whatever the user already built.
    Steps whose method isn't a process type this app models (AVAILABLE_PROCESSES) are
    skipped defensively - real archive data could contain types not covered here yet."""
    steps = _distinct_positions_in_order(cache.get_processing_steps(url, token, batch_id))
    sequence: list[ProcessInstance] = []
    for step in steps:
        process_type = resolve_process_type(step)
        if process_type is None or process_type == "Experiment Info":
            continue
        config = default_config_for(process_type)
        for key, count in infer_config_from_source_step(process_type, step).items():
            config[key] = max(config.get(key, 0), count)
        sequence.append(
            ProcessInstance(
                process_type=process_type, sequence_index=len(sequence) + 1, config=config
            )
        )
    return sequence


def clear_autofilled_value(spec: ProcessFieldSpec, sources: set[str]) -> None:
    """Clears value(s) whose provenance.source is in `sources`. Never touches a manually
    edited value - used to 'reset' prior autofill when a template/override is re-applied,
    per 'every autofilled field must always remain manually editable'."""
    if spec.varies:
        for sample_number in list(spec.per_sample_values):
            provenance = spec.per_sample_provenance.get(sample_number)
            if provenance is not None and provenance.source in sources:
                del spec.per_sample_values[sample_number]
                del spec.per_sample_provenance[sample_number]
        return
    if spec.provenance is not None and spec.provenance.source in sources:
        spec.value = None
        spec.provenance = None
        spec.is_outlier = False


def apply_whole_experiment_template(
    state: ExperimentState,
    url: str,
    token: str,
    cache: NomadSessionCache,
    batch_id: str,
    progress_callback=None,
) -> dict[int, int]:
    """Whole-experiment template scope: REPLACES the entire process sequence with one
    mirroring batch_id's own steps (see build_process_sequence_from_batch - one process
    per distinct sequence position, config sized to the source's actual data), then
    autofills every process from that same batch. This is a product decision, not an
    accident: replicating a whole experiment (a batch with 10 steps becomes a 10-step
    sequence), not just filling values into whatever the user had already built - so any
    processes/overrides/manual edits the user had before picking a template are
    discarded, same as re-picking a different template batch.

    `progress_callback`, if given, is called as progress_callback(done, total) after each
    process is autofilled - this is the only widget-adjacent hook in this module (still
    zero widget imports: it's a plain callable, e.g. gui_components wires it to a
    FloatProgress). This step is genuinely slow (one or more real HTTP calls per process,
    unlike a single-process override), so the caller can show real incremental progress
    instead of one static "working" message for the whole operation.

    Returns {sequence_index: fields_written} for every process that got at least one
    field written."""
    state.process_sequence = build_process_sequence_from_batch(url, token, cache, batch_id)
    rebuild_field_specs(state)
    written_by_process: dict[int, int] = {}
    total = len(state.process_sequence)
    for done, process in enumerate(state.process_sequence, start=1):
        occurrence = occurrence_index_for_process(state, process)
        written = autofill_process_from_batch(
            process, url, token, cache, batch_id, occurrence, provenance_source="batch_template"
        )
        if written:
            written_by_process[process.sequence_index] = written
        if progress_callback is not None:
            progress_callback(done, total)
    state.whole_experiment_template_batch_id = batch_id
    return written_by_process


def apply_process_override(
    state: ExperimentState,
    process: ProcessInstance,
    url: str,
    token: str,
    cache: NomadSessionCache,
    batch_id: str,
    occurrence: int | None = None,
) -> int:
    """Per-process override scope: clears this process's prior autofill (whole-template or
    a previous override - never manual edits) and refills from the new batch. Only this
    one process is touched; everything else stays tied to the whole-experiment template.

    `occurrence` defaults to this process's positional occurrence within the target
    sequence (occurrence_index_for_process); pass it explicitly to instead adopt a
    specific occurrence the user picked by hand (e.g. by material name, see
    list_process_occurrences) from a batch with multiple same-type steps.

    Before autofilling, widens this process's config to match how much data the source
    occurrence actually has (see expand_process_config_for_source) - otherwise values
    beyond the target's current config count (e.g. a 2nd solvent) would be silently
    dropped since there'd be no field_spec slot to write them into."""
    for spec in process.field_specs.values():
        clear_autofilled_value(spec, {"batch_template", "process_override"})
    if occurrence is None:
        occurrence = occurrence_index_for_process(state, process)
    expand_process_config_for_source(state, process, url, token, cache, batch_id, occurrence)
    written = autofill_process_from_batch(
        process, url, token, cache, batch_id, occurrence, provenance_source="process_override"
    )
    process.source_override_batch_id = batch_id
    return written


def clear_process_override(state: ExperimentState, process: ProcessInstance) -> None:
    """Removes the override: process_override-sourced values are cleared (never manual
    edits) and the process becomes tied to the whole-experiment template again. Does NOT
    automatically re-autofill from the whole template - call apply_whole_experiment_template
    again (or leave the now-empty fields for the next nudge pass) if that's desired."""
    for spec in process.field_specs.values():
        clear_autofilled_value(spec, {"process_override"})
    process.source_override_batch_id = None


def compute_field_distribution_for_occurrence(
    url: str,
    token: str,
    cache: NomadSessionCache,
    batch_id: str,
    process_type: str,
    field_key: str,
    occurrence: int = 0,
    exclude_occurrence: bool = False,
) -> list[float]:
    """Numeric values of field_key across every step sharing the SAME experimental-plan
    position as the given occurrence (i.e. across the real "variations" NOMAD recorded at
    that position - see the real file's per-Variation-group solvent ratios), not across
    every occurrence of process_type in the whole sequence.

    exclude_occurrence=True drops the occurrence's own step from the returned values -
    use this when testing whether THAT value is an outlier relative to the others.
    Including it (the default) suffers from statistical masking: with a self-inclusive
    population stdev and few points, one extreme value inflates its own stdev enough that
    it can never exceed a fixed z-score threshold (for n<=5 points the maximum possible
    self-inclusive z-score is under 2 regardless of how extreme the value is) - so
    is_outlier() would silently never fire on small, realistic sample counts."""
    field_paths = PROCESS_TYPE_FIELD_PATHS.get(process_type, {})
    path_info = field_paths.get(field_key)
    if path_info is None:
        return []
    paths, _ = path_info
    steps = steps_for_process_type(cache.get_processing_steps(url, token, batch_id), process_type)
    if occurrence >= len(steps):
        return []
    target_step = steps[occurrence]
    target_position = target_step.get("positon_in_experimental_plan")
    values = []
    for step in steps:
        if step.get("positon_in_experimental_plan") != target_position:
            continue
        if exclude_occurrence and step is target_step:
            continue
        value = _get_path_any(step, paths)
        if isinstance(value, int | float):
            values.append(float(value))
    return values


def is_outlier(value: Any, distribution: list[float], k: float = 2.0) -> bool:
    """Flags `value` if more than k standard deviations from the mean of `distribution`.
    Needs >=2 points to be meaningful; k=2.0 is a starting default - tune after real
    feedback, same as the nudge count/strategy."""
    if not isinstance(value, int | float) or len(distribution) < 2:
        return False
    mean = statistics.mean(distribution)
    std = statistics.pstdev(distribution)
    if std == 0:
        return False
    return abs(float(value) - mean) > k * std


# ---------------------------------------------------------------------------
# Varying-fields matrix + Variation column - experiment-wide (spans Experiment Info and
# every process), one checkbox per field ("this field varies"), matrix rows = samples
# (mothers; children always inherit). The Variation column is always freshly (re)computed
# from currently-checked fields, never autofilled/inherited, no-clobber into empty cells
# only - see the addendum in the task brief.
# ---------------------------------------------------------------------------


def iter_varying_fields(state: ExperimentState) -> list[tuple[str, ProcessFieldSpec]]:
    """(display_label, spec) for every field currently marked varies=True, in column
    order (Experiment Info first, then each process in sequence order). Excludes
    "Variation" itself, which is a computed output column, never a matrix input."""
    fields: list[tuple[str, ProcessFieldSpec]] = []
    for key, spec in state.experiment_info_fields.items():
        if key == "Variation":
            continue
        if spec.varies:
            fields.append((f"Experiment Info - {key}", spec))
    for process in state.process_sequence:
        for key, spec in process.field_specs.items():
            if spec.varies:
                fields.append((f"{process.sequence_index}: {process.process_type} - {key}", spec))
    return fields


def compute_variation_label(
    state: ExperimentState, sample_number: int, delimiter: str = "_"
) -> str:
    """Concatenates every checked-varying field's value for this sample, in column order,
    joined with a fixed delimiter (v1 - simple join, no per-field suffix/group-break)."""
    parts = [
        str(spec.per_sample_values.get(sample_number))
        for _label, spec in iter_varying_fields(state)
        if _is_filled(spec.per_sample_values.get(sample_number))
    ]
    return delimiter.join(parts)


def update_variation_column(state: ExperimentState, delimiter: str = "_") -> int:
    """No-clobber (re)computation of the Variation column for every sample - only writes
    into currently-empty per-sample slots. Returns how many slots were written. Safe to
    call after every matrix edit; already-set Variation values (manual or previously
    computed) are never touched, per the no-clobber rule."""
    variation_spec = state.experiment_info_fields.get("Variation")
    if variation_spec is None:
        return 0
    variation_spec.varies = True
    written = 0
    for sample_number in state.sample_numbers():
        if _is_filled(variation_spec.per_sample_values.get(sample_number)):
            continue
        label = compute_variation_label(state, sample_number, delimiter)
        if not label:
            continue
        variation_spec.per_sample_values[sample_number] = label
        variation_spec.per_sample_provenance[sample_number] = FieldProvenance(source="computed")
        written += 1
    return written


# ---------------------------------------------------------------------------
# Material-gated progress bar. A process only counts toward the denominator if it is not
# material-gated, or is material-gated and has "Material name" filled - see
# ProcessInstance.counts_toward_progress(). Field-level counting is done in alias-grouped
# "progress units" (alias_config.resolve_progress_units) so config-dependent field
# renaming (e.g. single- vs multi-step Spin Coating rotation fields) doesn't double-count.
# Scoped to process_sequence only - Experiment Info fields are not counted (not mentioned
# in the material-gating spec, and typically filled at experiment setup, not per-process).
# ---------------------------------------------------------------------------


def compute_process_progress(
    process: ProcessInstance, alias_groups: list[dict] | None = None
) -> tuple[int, int]:
    """(filled_units, total_units) for one process, after alias-group merging. (0, 0) if
    the process is material-gated and has no material yet (excluded entirely, per the
    material-gating rule - matches the real NOMAD parser, which skips such processes).
    Fields with required_for_progress=False are dropped before unit resolution, so they
    never appear in the denominator OR numerator - the user-facing "this field doesn't
    matter" opt-out (see set_field_required_for_progress)."""
    if not process.counts_toward_progress():
        return (0, 0)
    field_keys = [key for key, spec in process.field_specs.items() if spec.required_for_progress]
    units = resolve_progress_units(process.process_type, field_keys, alias_groups)
    filled = sum(1 for unit in units if any(process.field_specs[key].is_filled() for key in unit))
    return (filled, len(units))


def compute_experiment_progress(
    state: ExperimentState, alias_groups: list[dict] | None = None
) -> tuple[int, int]:
    """(filled_units, total_units) summed across every process in the sequence."""
    total_filled = 0
    total_units = 0
    for process in state.process_sequence:
        filled, total = compute_process_progress(process, alias_groups)
        total_filled += filled
        total_units += total
    return (total_filled, total_units)


def progress_band(filled: int, total: int) -> str:
    """Coarse completion band for ProgressBarWidget's color-coding: 'red' at 1/3 full or
    less, 'yellow' below 2/3, 'green' at 90% or more, 'blue' in between. An empty
    denominator (nothing counts toward progress yet) is treated as 'red'."""
    if total <= 0:
        return "red"
    ratio = filled / total
    if ratio <= 1 / 3:
        return "red"
    if ratio < 2 / 3:
        return "yellow"
    if ratio < 0.9:
        return "blue"
    return "green"


def compute_experiment_info_progress(state: ExperimentState) -> tuple[int, int]:
    """(filled, total) for state.experiment_info_fields, excluding computed-only keys
    (Variation/Nomad ID/Sample), pixel fields (edited via SampleSetupPanel/pixel_fields
    instead), and fields the user has marked required_for_progress=False - the same
    filter ExperimentInfoPanel uses to decide what it renders (plus the opt-out). Used to
    show a completion percentage next to the Experiment Info row's title."""
    relevant = [
        spec
        for key, spec in state.experiment_info_fields.items()
        if key not in EXPERIMENT_INFO_COMPUTED_KEYS
        and key not in PIXEL_FIELD_KEYS
        and spec.required_for_progress
    ]
    filled = sum(1 for spec in relevant if spec.is_filled())
    return (filled, len(relevant))


# ---------------------------------------------------------------------------
# Nudge queue - guided popup flow prioritizing empty/unfilled fields over already-filled
# ones (replaces plain random spot-checking), ending with a summary of remaining gaps.
# Outlier flagging on filled values (ProcessFieldSpec.is_outlier, set during autofill
# above) still applies separately and is queued after the missing-field items.
#
# Scoped to non-varying fields only: a varying field's "value" is per-sample
# (per_sample_values), which doesn't map onto this queue's single-value edit widget: the
# VaryingFieldsMatrix is already the visible, editable surface for those. Unlike the
# progress bar, this queue does NOT exclude material-gated processes with an empty
# material field - the point of nudging is to help the user close exactly that kind of
# gap, not hide it because the progress bar isn't counting it yet.
# ---------------------------------------------------------------------------


class NudgeItem(BaseModel):
    sequence_index: int
    process_type: str
    field_key: str
    kind: Literal["missing", "outlier"]


def build_nudge_queue(state: ExperimentState, max_items: int | None = None) -> list[NudgeItem]:
    """Missing fields first, grouped by process with the worst gaps (most missing fields)
    first, then outlier-flagged filled fields in sequence order. max_items caps the total
    queue length (None = no cap) - kept as a parameter so the strategy is easy to retune,
    per the product owner's expectation that this gets tuned after real feedback. Fields
    with required_for_progress=False are skipped entirely - marking a field "not
    required" means don't nag about it either, not just don't count it."""
    processes_with_missing: list[tuple[ProcessInstance, list[str]]] = []
    for process in state.process_sequence:
        missing_keys = [
            key
            for key, spec in process.field_specs.items()
            if not spec.varies and not spec.is_filled() and spec.required_for_progress
        ]
        if missing_keys:
            processes_with_missing.append((process, missing_keys))
    processes_with_missing.sort(key=lambda pair: len(pair[1]), reverse=True)

    missing_items = [
        NudgeItem(
            sequence_index=process.sequence_index,
            process_type=process.process_type,
            field_key=field_key,
            kind="missing",
        )
        for process, missing_keys in processes_with_missing
        for field_key in missing_keys
    ]

    outlier_items = [
        NudgeItem(
            sequence_index=process.sequence_index,
            process_type=process.process_type,
            field_key=field_key,
            kind="outlier",
        )
        for process in state.process_sequence
        for field_key, spec in process.field_specs.items()
        if not spec.varies and spec.is_outlier and spec.is_filled() and spec.required_for_progress
    ]

    queue = missing_items + outlier_items
    return queue if max_items is None else queue[:max_items]


def build_missing_fields_summary(state: ExperimentState) -> list[tuple[ProcessInstance, int]]:
    """Every process that still has missing REQUIRED fields (varying or not - uses the
    same is_filled() the progress bar uses; excludes required_for_progress=False fields,
    same as the progress bar), with a count of how many. Always the LAST popup shown in
    the nudge flow, regardless of queue length."""
    summary = []
    for process in state.process_sequence:
        missing_count = sum(
            1
            for spec in process.field_specs.values()
            if not spec.is_filled() and spec.required_for_progress
        )
        if missing_count:
            summary.append((process, missing_count))
    return summary


# ---------------------------------------------------------------------------
# Excel finalization - reuses Excel_creator's ExperimentExcelBuilder exactly for the
# header/guide/citation sheets, then smart_databaser writes real data rows (mother +
# child) itself, never Excel_creator's single fabricated is_testing row.
# ---------------------------------------------------------------------------

PARENT_ID_COLUMN_KEY = "Parent ID"


def enumerate_sample_rows(state: ExperimentState) -> list[tuple[int, int | None]]:
    """(sample_number, child_index) pairs in output row order: each sample's mother row
    (child_index=None) first, then its children 1..child_count, per sample in
    state.samples order."""
    rows: list[tuple[int, int | None]] = []
    for sample in state.samples:
        rows.append((sample.sample_number, None))
        rows.extend(
            (sample.sample_number, child_index) for child_index in range(1, sample.child_count + 1)
        )
    return rows


def compute_nomad_id(state: ExperimentState, sample_number: int, child_index: int | None) -> str:
    """Best-effort ID scheme built only from fields this model tracks (Project_Name,
    Batch, Subbatch, Sample) plus a '_C-{child_index}' suffix for child rows.

    UNVERIFIED against the real NOMAD parser's exact expected format: the one real
    historical file inspected during planning (20260603_Batch2028.xlsx) embeds an extra
    numeric segment ("HZB_CsFA_8_2028_1_C-0") that doesn't map to any field this app
    currently tracks (Subbatch was empty in that row, so the "8" is neither Batch nor
    Subbatch - likely some upload/experiment number not modeled here). Confirm/adjust
    this scheme with a Data Steward before relying on it in production, same as the
    Parent ID column-position assumption."""
    parts = []
    for key in ("Project_Name", "Batch", "Subbatch"):
        spec = state.experiment_info_fields.get(key)
        if spec is not None and _is_filled(spec.value):
            parts.append(str(spec.value))
    parts.append(str(sample_number))
    nomad_id = "HZB_" + "_".join(parts)
    if child_index is not None:
        nomad_id += f"_C-{child_index}"
    return nomad_id


def _resolve_experiment_info_cell(
    state: ExperimentState, field_key: str, sample_number: int
) -> Any:
    spec = state.experiment_info_fields.get(field_key)
    if spec is None:
        return None
    if spec.varies:
        return spec.per_sample_values.get(sample_number)
    return spec.value


def _resolve_pixel_cell(
    state: ExperimentState, field_key: str, sample_number: int, child_index: int | None
) -> Any:
    """Pixel-specific fields are blank on the mother row - confirmed against the real
    file, where 'Number of pixels'/'Pixel area' are empty for the whole-substrate row and
    only populated on individual diced-pixel child rows."""
    if child_index is None:
        return None
    spec = state.pixel_fields.get(field_key)
    if spec is None:
        return None
    if spec.varies:
        return spec.per_child_values.get((sample_number, child_index))
    return spec.value


def _resolve_process_cell(
    state: ExperimentState, sequence_index: int, field_key: str, sample_number: int
) -> Any:
    try:
        process = state.get_process(sequence_index)
    except KeyError:
        return None
    spec = process.field_specs.get(field_key)
    if spec is None:
        return None
    if spec.varies:
        return spec.per_sample_values.get(sample_number)
    return spec.value


def resolve_cell_value(
    state: ExperimentState,
    sequence_index: int,
    field_key: str,
    sample_number: int,
    child_index: int | None,
) -> Any:
    """A varying field's per-sample value is shared by a sample's mother and every child
    row (children inherit the same process-column values as their mother, per the real
    file's structure) - only pixel fields differ per child row."""
    if sequence_index == 0:
        if field_key in PIXEL_FIELD_KEYS:
            return _resolve_pixel_cell(state, field_key, sample_number, child_index)
        return _resolve_experiment_info_cell(state, field_key, sample_number)
    return _resolve_process_cell(state, sequence_index, field_key, sample_number)


def append_parent_id_column(worksheet: Worksheet) -> int:
    """Appends 'Parent ID' at the end of the worksheet rather than splicing it into the
    Experiment Info block, to avoid disturbing sheet_experiment.py's existing data
    validations/merged ranges (which reference fixed column letters computed at
    generation time). Assumes NOMAD's parser matches columns by header text, not
    position - worth a sanity check with a Data Steward once this is live."""
    column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=column_index, value="Experiment Info")
    worksheet.cell(row=2, column=column_index, value=PARENT_ID_COLUMN_KEY)
    return column_index


def generate_full_workbook(state: ExperimentState) -> Workbook:
    """Full 3-sheet workbook (Experiment Data + Data Entry Guide + How to Cite) via
    Excel_creator's own ExperimentExcelBuilder, then real mother+child data rows written
    by smart_databaser itself. Nomad ID, "Sample", and Parent ID are always freshly
    computed at write time from sample_number/child_index (never sourced from
    field_specs, never no-clobbered) - they're derived identifiers, not user data, so
    requiring the user to manually mark them varying and re-enter the same numbers per
    row would be redundant."""
    builder = ExperimentExcelBuilder(process_sequence_to_dicts(state), is_testing=False)
    builder.build_excel()
    workbook = builder.workbook
    worksheet = workbook["Experiment Data"]
    column_map = build_column_map(worksheet)
    parent_id_col = append_parent_id_column(worksheet)
    nomad_id_col = column_map.get((0, "Nomad ID"))
    sample_col = column_map.get((0, "Sample"))

    mother_nomad_ids: dict[int, str] = {}
    for row_offset, (sample_number, child_index) in enumerate(enumerate_sample_rows(state)):
        row = 3 + row_offset
        nomad_id = compute_nomad_id(state, sample_number, child_index)
        if child_index is None:
            mother_nomad_ids[sample_number] = nomad_id

        for (sequence_index, field_key), col in column_map.items():
            if sequence_index == 0 and field_key in ("Nomad ID", "Sample"):
                continue
            value = resolve_cell_value(state, sequence_index, field_key, sample_number, child_index)
            if value is not None:
                worksheet.cell(row=row, column=col, value=value)

        if nomad_id_col is not None:
            worksheet.cell(row=row, column=nomad_id_col, value=nomad_id)
        if sample_col is not None:
            worksheet.cell(row=row, column=sample_col, value=sample_number)
        if child_index is not None:
            worksheet.cell(row=row, column=parent_id_col, value=mother_nomad_ids.get(sample_number))

    return workbook


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_experiment_filename() -> str:
    """Reuses ExperimentExcelBuilder.save()'s exact filename convention
    (YYYYMMDD_experiment_file.xlsx) - not File_Uploader's per-sample renaming scheme,
    since NOMAD's parser for these files is keyed on structure/content, and the
    date/time in the name is the de facto creation timestamp."""
    return f"{datetime.now().strftime('%Y%m%d')}_experiment_file.xlsx"


# ---------------------------------------------------------------------------
# Upload mechanism - distinct decision point, per the plan: pushes the generated Excel's
# raw bytes into a user-selected, already-existing NOMAD upload (the manual "create the
# upload in the NOMAD web GUI first" step is kept - this app never auto-creates an
# upload). Adapted from File_Uploader.gui_components.upload_files_for_samples /
# _process_upload (apps/File_Uploader/gui_components.py:506-609), but: a single file (no
# zip - one experiment Excel, not per-sample files), the correct content-type
# (File_Uploader hardcodes "application/json" for its zip - a pre-existing quirk not
# carried over here), and explicit raise_for_status() on every call (File_Uploader's
# _process_upload has none, so a failed process/poll call there fails silently or hangs -
# not carried over either). The target upload_id comes from
# hysprint_utils.api_calls.get_all_uploads (via NomadSessionCache.get_uploads), not from
# an existing sample lookup like File_Uploader's get_nomad_ids_of_entry.
#
# MUST be live-verified against a disposable test upload (see
# tests/live/test_smart_databaser_upload.py) before this is trusted against a real
# upload - it has not been exercised against the real API as of this writing.
# ---------------------------------------------------------------------------

UPLOAD_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def upload_experiment_excel(
    url: str,
    token: str,
    upload_id: str,
    filename: str,
    excel_bytes: bytes,
    poll_interval_seconds: float = 2.0,
    max_poll_seconds: float = 120.0,
) -> None:
    """PUTs the Excel's raw bytes into upload_id's /raw/ endpoint, triggers processing,
    then polls until data.process_running is False. Raises requests.HTTPError on any
    failed call, or TimeoutError if processing doesn't finish within max_poll_seconds."""
    put_response = requests.put(
        f"{url}/uploads/{upload_id}/raw/",
        data={"wait_for_processing": False},
        headers={"Authorization": f"Bearer {token}"},
        files={"file": (filename, excel_bytes, UPLOAD_MIME_TYPE)},
    )
    put_response.raise_for_status()

    process_response = requests.post(
        f"{url}/uploads/{upload_id}/action/process",
        headers={"Authorization": f"Bearer {token}"},
    )
    process_response.raise_for_status()

    elapsed = 0.0
    while elapsed < max_poll_seconds:
        time.sleep(poll_interval_seconds)
        elapsed += poll_interval_seconds
        status_response = requests.get(
            f"{url}/uploads/{upload_id}", headers={"Authorization": f"Bearer {token}"}
        )
        status_response.raise_for_status()
        if not status_response.json()["data"]["process_running"]:
            return

    raise TimeoutError(f"Upload {upload_id} processing did not finish within {max_poll_seconds}s")

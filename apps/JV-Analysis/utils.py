"""
Utility Functions Module
Contains Excel export, file operations, and other utility functions.
Extracted from main.py for better organization.
"""

__author__ = "Edgar Nandayapa"
__institution__ = "Helmholtz-Zentrum Berlin"
__created__ = "August 2025"

import os
from datetime import date

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def dated_filename(filename):
    """Prefix a filename with today's date (YYYY-MM-DD) for easy identification."""
    return f"{date.today().isoformat()}_{filename}"


# Summary-table columns storing signed current density (stored negative internally
# to match the raw JV-curve sign convention; flipped positive on export for readability).
CURRENT_DENSITY_SUMMARY_COLUMNS = ["Jsc(mA/cm2)", "J_mpp(mA/cm2)"]

CURRENT_DENSITY_FLIP_NOTE = (
    "Note: current density values (Jsc, J_mpp, Current Density) have been multiplied "
    "by -1 from their internal storage so the reported magnitude is positive. The "
    "app's plots are unaffected and keep the original sign convention."
)

# Non-data columns in the long-format curves table (see data_manager.py's columns_cur).
_CURVE_META_COLUMNS = {
    "index",
    "sample",
    "batch",
    "condition",
    "variable",
    "cell",
    "direction",
    "ilum",
    "sample_id",
    "status",
}


def flip_current_density_sign(df, columns=CURRENT_DENSITY_SUMMARY_COLUMNS):
    """Return a copy of df with the given current-density columns sign-flipped."""
    df = df.copy()
    for col in columns:
        if col in df.columns:
            df[col] = df[col] * -1
    return df


def flip_current_density_curve_rows(df, variable_col="variable", target="Current Density(mA/cm2)"):
    """Return a copy of df with curve-point columns sign-flipped for current-density rows."""
    df = df.copy()
    if df.empty or variable_col not in df.columns:
        return df
    mask = df[variable_col] == target
    if not mask.any():
        return df
    value_cols = [c for c in df.columns if c not in _CURVE_META_COLUMNS]
    df.loc[mask, value_cols] = df.loc[mask, value_cols] * -1
    return df


def save_full_data_frame(data):
    """
    Create and return an Excel workbook with the full dataframe.
    Simplified version that just creates a workbook without saving to file.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Add main data sheet
    ws = wb.create_sheet(title="All_data")
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)

    return wb


def save_combined_excel_data(path, wb, data, filtered_info, var_x, name_y, var_y, other_df):
    """Save combined data to Excel workbook with multiple sheets"""
    trash, filters = filtered_info

    # Create sheet name based on variables
    sheet_title = f"{var_y}-by-{var_x}"

    # Check if the sheet already exists and remove it
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    # Insert header
    ws.append([f"Contents of boxplot for {var_y} by {var_x}"])
    if name_y in CURRENT_DENSITY_SUMMARY_COLUMNS:
        ws.append([CURRENT_DENSITY_FLIP_NOTE])
    ws.append([])  # Empty row

    # Process and append main data
    combined_data = data.copy()
    combined_data["_index"] = combined_data.groupby(var_x).cumcount()
    pivot_table = combined_data.pivot_table(
        index="_index", columns=var_x, values=name_y, aggfunc="mean"
    )

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        ws.append(r)

    # Add statistical summary
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="Statistical summary")
    ws.append([])

    for r in dataframe_to_rows(other_df.T, index=True, header=True):
        ws.append(r)

    # Add filtered data section
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="This is the filtered data")
    ws.append([])

    if not trash.empty:
        combined_trash = trash.copy()
        combined_trash["_index"] = combined_trash.groupby(var_x).cumcount()
        pivot_table_trash = combined_trash.pivot_table(
            index="_index", columns=var_x, values=name_y, aggfunc="mean"
        )

        for r in dataframe_to_rows(pivot_table_trash, index=True, header=True):
            ws.append(r)

    # Add filter information
    next_row = ws.max_row + 3
    filter_words = ["Only data within these limits is shown:"] + filters
    for cc, strings in enumerate(filter_words):
        ws.cell(row=next_row + cc, column=1, value=strings)

    return wb


def is_running_in_jupyter():
    """Check if code is running in Jupyter notebook"""
    try:
        from IPython import get_ipython

        return get_ipython() is not None
    except ImportError:
        return False


def create_new_results_folder(path):
    """Create a results folder if it doesn't exist"""
    folder_path = os.path.join(path, "Results")
    try:
        os.makedirs(folder_path, exist_ok=True)
    except Exception as e:
        print(f"Warning: Could not create results folder: {e}")
        return path
    return folder_path


def clean_filename(filename):
    """Clean filename for safe saving"""
    import re

    # Remove invalid characters for filenames
    filename = re.sub(r'[<>:"/\\|?*]', "_", filename)
    return filename
